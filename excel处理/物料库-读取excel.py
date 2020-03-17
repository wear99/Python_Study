# 读取指定的excel文件,将物料属性以list形式存储，并写入json文件。
# 增加读取已有文件功能，编码存在则不写入
# -*-coding=utf-8-*-

import openpyxl as xl
import json
from datetime import datetime

# 根据excel 行读取到列表中
def add_item(n):
    code = std(n[1].value)
    draw = std(n[4].value)
    name = std(n[2].value)
    data = std(n[5].value)

    if code and code != '-':
        item = [code, draw, name, data]
        if code not in all_code_dict:
            all_code.append(item)

def std(s):
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    elif s:
        return str(s).upper().strip()
    else:
        return '-'

# 查找旧编码,并在编码前加标记 *
def old_item_check():
    all_code.sort(key=lambda x: x[0])
    # 物料按编码排序,对任一编码和下一个编码进行比对,如果除了最后一位相同,且比较小,则认为是旧编码
    for x in range(len(all_code) - 1):
        
        a = all_code[x][0]
        if a[2] == 'R':
            if a[-1] == 'P':
                a = a.replace('P', '')
            b = all_code[x + 1][0]

            if b[-1] == 'P':
                all_code[x][0] = all_code[x][0]+' old'
            elif a == b:
                all_code[x][0] = all_code[x][0]+' rpt'
            elif a[:-1] == b[:-1]:
                if a[-1] <= b[-1]:
                    all_code[x][0] = all_code[x][0]+' old'

all_code = []       # 所有编码的库文件，读取完毕会写入到文件中

all_code_dict = {}    # 所以编码的字典文件，便于查找重复项
#all_code_rpt = []     # 添加时重复的编码会被加入进来

filename = ['折叠机加工件新编码.xlsx', '各种采购件新编码.xlsx', '滚筒烫平机加工件新编码.xlsx','OEM&集成系统&能效系统加工件新编码.xlsx']
patch = 'D:\\work\\python\\excel处理\\excel\\'
file = [patch + x for x in filename]
  
# 用于读取已有json文件，并生成一个字典文件,用来快速判断编码是否存在
try:
    with open('all_code.json', 'r', encoding='utf-8') as fr:
        all_code = json.load(fr)
    for item in all_code:
        all_code_dict[item[0]]=' '
except:
    print('现有文件不存在')

for wbook in file:
    wb = xl.load_workbook(wbook, read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for m in ws.iter_rows(min_row=2, max_col=6):
            add_item(m)

old_item_check()

print(len(all_code))

# indent 有缩进，后一个是不用ascii编码，就可以在文件中显示成中文
# 把读取的物料表写入json文件,方便快速读取
with open('all_code.json', 'w', encoding='utf-8') as fw:
    json.dump(all_code, fw, indent=4, ensure_ascii=False)

