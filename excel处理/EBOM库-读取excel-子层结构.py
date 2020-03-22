# -*-coding=utf-8-*-
# 整个bom库为字典结构，每个物料编码为key,其第1子层物料的编码和数量存储到该物料的值

import openpyxl as xl
import json

def add_item(wsheet, lv_c, code_c, num_c, start_r, root,mode='R'):
    all_bom_1={}
    all_bom_1[root] = []
    lv_code = {0:root,}
    ex_code_lv=10
    for n in wsheet.iter_rows(min_row=start_r, max_col=num_c + 1):
        code = n[code_c].value
        lv = len(n[lv_c].value)
        num = float(n[num_c].value)
        if lv > ex_code_lv:
            continue
        else:
            all_bom_1[lv_code[lv - 1]].append([code, num])

        if code in all_bom_1:
            ex_code_lv = lv
        else:
            lv_code[lv] = code
            all_bom_1[code] = []
            ex_code_lv = 10

    # R 模式: 对于已经存在的编码,全部跳过.
    # A 模式: 对于已存在的,如果子项为空,则重新写入.
    # W 模式: 对于已存在的,把原来的清空,按新导入重新添加子项

    for key,item in all_bom_1.items():
        if mode == 'W' or key not in all_bom_code:
            all_bom_code[key] = item
        elif mode == 'A':
            if all_bom_code[key] == "":
                all_bom_code[key] = item
        else:
            pass

def read_date(date):
    try:
        with open(date, 'r', encoding='utf-8') as fb:
            return json.load(fb)
    except:
        print('文件不存在，重新建立')
        return {'index':{}}

def read_excel(filelist):
    def get_root(wsheet):
        m = 0
        for n in wsheet.rows:
            m += 1
            if m > 10:
                break
            if n[0].value.upper() in all_root:
                return n[0].value,n[1].value
        return '', ''

    def get_col(wsheet):
        start_r = 1
        lv_c, code_c, num_c = 100, 100, 100

        for n in wsheet.rows:
            start_r += 1
            if start_r > 10:
                break
            for col in range(0, len(n)):
                if n[col].value and ('级别' in str(n[col].value) or '层次' in str(n[col].value)):
                    lv_c = col
                if n[col].value and '编码' in str(n[col].value):
                    code_c = col
                if n[col].value and ('用量' in str(n[col].value) or '数量' in str(n[col].value)):
                    num_c = col

                if lv_c!=100 and code_c!=100 and num_c!=100:
                    return lv_c, code_c, num_c, start_r

    for wbook in filelist:
        wb = xl.load_workbook(wbook, read_only=True)
        for wsname in wb.sheetnames:
            ws = wb[wsname]
            # 查找sheet表内的产品型号
            root, rootname = get_root(ws)
            if not root:
                print('{0} 文件的{1} 工作表中未找到产品代码'.format(wbook, wsname))
                continue
            elif root in all_bom_code['index']:
                print('{0} 文件的{1} 工作表中的产品代码{2}在BOM库中已存在'.format(wbook, wsname,root))
                continue

            # 查找层次,编码,数量 对应的列数,及起始行数
            lv_c, code_c, num_c, start_r = get_col(ws)
            if not start_r:
                print('{0} 文件的{1} 工作表中未找到物料属性表头，跳过'.format(wbook, wsname))
                continue

            # 添加物料到all_bom
            all_bom_code['index'][root]=rootname
            add_item(ws, lv_c, code_c, num_c, start_r,root,mode='A')

# 所有产品的代码，不包含在内的将不被读取
all_root = {
    'C07-0024': '卓越CEF-33-1L-FS折叠堆码机',
    'C07-0026': '卓越CEF-33-1L-F-FS折叠机堆码机',
    'C08-0005': 'CSJ-150穿梭机',
    'C07-0020': '卓越CETF-20毛巾折叠机',
    'C07-0010': '卓越CEF-30-2L折叠机'
}

filename = ['ZD折叠机.xlsx']
patch = 'D:\\work\\python\\excel处理\\excel\\'
file = [patch + x for x in filename]

all_bom_code = read_date('all_bom_code.json')

read_excel(file)

print('已读取的BOM数量：',len(all_bom_code)-1)
for x in all_bom_code['index']:
    print(x)

with open('all_bom_code.json', 'w', encoding='utf-8') as f:
    json.dump(all_bom_code, f, indent=4, ensure_ascii=False)
