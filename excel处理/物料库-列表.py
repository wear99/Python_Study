# 读取指定的excel文件,将物料属性以list形式存储，并写入json文件
# -*-coding=utf-8-*-

import openpyxl as xl
import json
import time
from datetime import datetime

# 根据excel 行读取到列表中
def add_item(n):
    code = std(n[1].value)
    draw = std(n[4].value)
    name = std(n[2].value)
    data = std(n[5].value)

    if code:
        x = [code, draw, name, data]
        if code:
            all_code.append(x)


def std(s):
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    elif s:
        return str(s).upper().strip()


def find(s):
    rst = []
    for item in all_code_latest:
        for m in item:
            if m and s in m:
                rst.append(item)
    if rst:
        #rst.sort(key=lambda x: x[0])
        print('共查找到物料: %d' % len(rst))
        for key in rst:
            print('%-12s  %-25s  %-40s  %-10s' %
                  (key[0], key[1], key[2], key[3]))
    else:
        print('结果不存在！')

# 查找旧编码,并在编码前加标记 *
def old_item():
    for x in range(len(all_code)-1):

        #for y in all_code[x + 1:]:
        a = all_code[x][0]
        b = all_code[x + 1][0]
        a1 = all_code[x][0][:-1]
        b1 = all_code[x + 1][0][:-1]
        if a1 == b1 and a <= b:
            all_code[x][0] = '* ' + all_code[x][0]
            all_code_old.append(all_code[x])

        else:
            all_code_latest.append(all_code[x])
    all_code_latest.append(all_code[-1])

all_code = []
all_code_latest = []
all_code_old = []

开始时间 = time.time()
file = ['折叠机加工件新编码.xlsx', '各种采购件新编码.xlsx', '滚筒烫平机加工件新编码.xlsx']
excel = []
for x in file:
    excel.append('E:\\Python Study\\excel处理\\xlsx\\' + str(x))

for x in excel:
    wb = xl.load_workbook(x, read_only=True)

    for n in wb.sheetnames:
        ws = wb[n]
        for m in ws.iter_rows(min_row=2, max_col=6):
            add_item(m)

all_code.sort(key=lambda x: x[0])

old_item()

结束时间 = time.time()
print('读取完毕，执行时间为 %f 秒' % (结束时间 - 开始时间))

print(len(all_code))
print(len(all_code_old))
print(len(all_code_latest))

#print((len(item_rep)))
'''
while True:
    x = ''
    x = input('请输入要查询的物料编码或图号：\n').upper().strip()
    if x == 'q' or x == 'Q':
        break
    elif x:
        find(x)
'''
# indent 有缩进，后一个是不用ascii编码，就可以在文件中显示成中文
# 把读取的物料表写入json文件,方便快速读取
with open('all_code.json', 'w', encoding='utf-8') as f:
    json.dump(all_code, f, indent=4, ensure_ascii=False)
with open('all_code_old.json', 'w', encoding='utf-8') as g:
    json.dump(all_code_old, g, indent=4, ensure_ascii=False)
