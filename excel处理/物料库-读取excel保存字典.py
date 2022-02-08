# 读取指定的excel文件,将物料属性以list形式存储，并写入json文件。
# 增加读取已有文件功能，编码存在则不写入
# 
# -*-coding=utf-8-*-

import openpyxl as xl
import json
from datetime import datetime

# 根据excel 行读取到列表中
def add_item(n):
    code = std(n[1].value)
    draw = std(n[4].value)
    name = std(n[2].value)
    data = std(n[5].value)

    if code and code != '-':        
        if code in all_code:
            all_code_rpt[code] = [all_code[code], [code, draw, name, data]]
        else:
            all_code[code]=[code, draw, name, data]

def std(s):
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    elif s:
        return str(s).upper().strip()
    else:
        return '-'

# 查找旧编码,并在编码前加标记 *
def old_item_check():
    # 取出字典中所有的key,组成list，然后排序进行比较
    keys=list(all_code.keys())
    keys.sort()
    # 物料按编码排序,对任一编码和下一个编码进行比对,如果除了最后一位相同,且比较小,则认为是旧编码
    for x,a in enumerate(keys):
        if a[2] == 'R':
            if a[-1] == 'P':
                a = a.replace('P', '')
            b = keys[x + 1]
            if b[-1] == 'P':
                all_code[a][0] += ' old'            
            elif a[:-1] == b[:-1] and a[-1] <= b[-1]:
                all_code[a][0] +=' old'
        if x == len(keys) - 2:
            break

all_code = {}       # 所有编码的库文件，读取完毕会写入到文件中
all_code_rpt = {}    # 所以编码的字典文件，便于查找重复项

filename = ['折叠机加工件新编码.xlsx', '各种采购件新编码.xlsx', '滚筒烫平机加工件新编码.xlsx','OEM&集成系统&能效系统加工件新编码.xlsx']
patch = 'D:\\work\\python\\excel处理\\excel\\'
file = [patch + x for x in filename]

# 用于读取已有json文件
try:
    with open('all_code.json', 'r', encoding='utf-8') as fr:
        all_code = json.load(fr)
except:
    print('现有文件不存在')

for wbook in file:
    wb = xl.load_workbook(wbook, read_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for m in ws.iter_rows(min_row=2, max_col=6):
            add_item(m)

old_item_check()

print(len(all_code))

# indent 有缩进，后一个是不用ascii编码，就可以在文件中显示成中文
# 把读取的物料表写入json文件,方便快速读取
with open('all_code.json', 'w', encoding='utf-8') as fw:
    json.dump(all_code, fw, indent=4, ensure_ascii=False)

