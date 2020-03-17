# -*-coding=utf-8-*-
# 按bom表原样添加,产品码前层次为'*'. 整个bom库为字典结构, 每个产品代码为key,结构为值.

import openpyxl as xl
import json

def get_root(wsheet):
    m = 0
    for n in wsheet.rows:
        m += 1
        if m > 10:
            break
        if n[0].value.upper() in all_root:
            return n[0].value,n[1].value
    return '', ''

def get_col(wsheet):
    start_r = 1
    lv_c, code_c, num_c = 0, 0, 0

    for n in wsheet.rows:
        start_r += 1
        if start_r > 10:
            break
        for col in range(0, len(n)):
            if n[col].value and ('级别' in str(n[col].value) or '层次' in str(n[col].value)):
                lv_c = str(col)
            if n[col].value and '编码' in str(n[col].value):
                code_c = col
            if n[col].value and ('用量' in str(n[col].value) or '数量' in str(n[col].value)):
                num_c = col
            if lv_c and code_c and num_c:
                return int(lv_c), code_c, num_c, start_r

def add_item(wsheet):
    def std(s):
        if s:
            return str(s).upper().strip()

    for n in wsheet.iter_rows(min_row=start_row, max_col=num_col+1):
        bom_lv = n[lv_col].value + '+'
        code = std(n[code_col].value)
        #fu_code = get_fu(n)
        num = float(std(n[num_col].value))
        all_bom[root].append([bom_lv, code, num])

filename = ['ZD折叠机.xlsx']
patch = 'D:\\work\\python\\excel处理\\excel\\'
file = [patch + x for x in filename]

# 所有产品的代码，不包含在内的将不被读取
all_root = {
            'C07-0024':'卓越CEF-33-1L-FS折叠堆码机',
            'C07-0026':'卓越CEF-33-1L-F-FS折叠机堆码机',
            'C08-0005':'CSJ-150穿梭机',
            'C07-0020':'卓越CETF-20毛巾折叠机',
            'C07-0010':'卓越CEF-30-2L折叠机',
            }

# 尝试读取已存在的all_bom文件，并根据文件内的产品型号删除产品型号列表
try:
    with open('all_bom.json', 'r', encoding='utf-8') as fb:
        all_bom = json.load(fb)
    for root in all_bom:
        if root in all_root:
            del all_root[root]
except:
    all_bom = {'index':[]}
    print('文件不存在，重新建立')

for wbook in file:
    if not all_root:   #如果all_root为空，就不必执行后续
        break
    wb = xl.load_workbook(wbook, read_only=True)
    for wsname in wb.sheetnames:
        ws = wb[wsname]
        # 查找sheet表内的产品型号
        root, rootname = get_root(ws)       
        if root and root not in all_bom:
            pass
        else:
            print('{0} 文件的{1} 工作表中未找到产品或已存在'.format(wbook, wsname))
            continue

        # 查找层次,编码,数量 对应的列数,及起始行数
        lv_col, code_col, num_col, start_row = get_col(ws)
        if str(lv_col) and code_col and num_col and start_row:
            pass
        else:
            print('{0} 文件的{1} 工作表中未找到物料属性表头，跳过'.format(wbook, wsname))
            continue
        # 添加物料到all_bom 
        del all_root[root]
        all_bom['index'].append(['*',root+' '+ rootname,1])
        all_bom[root] = [['*', root, 1]]       
        add_item(ws)

print('已读取的BOM数量：',len(all_bom))
for x in all_bom['index']:
    print(x[1])

with open('all_bom.json', 'w', encoding='utf-8') as f:
    json.dump(all_bom, f, indent=4, ensure_ascii=False)
