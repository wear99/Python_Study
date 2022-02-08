import openpyxl as xl
from datetime import datetime

#wb = xl.load_workbook(r'D:\work\python\excel处理\excel\GPF-33 小批-高速折叠机试制清单EBOM.xlsx')

def read_design_BOM(file,sheet=''):  # 读取excel的bom文件,格式为导出样式,层次为+++
    # 先判断属性列和root,再读取并按统一格式生成列表:(任何列都允许不存在,由后续程序进行判断)
    # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    # 第一个是属性头
    def get_col(wsheet):
        col.clear()
        title = {
            '级别': 'lv',
            '层次': 'lv',
            '序号': 'lv',
            '编码': 'code',
            '子件编码':'code',
            '图号': 'draw',
            '代号': 'draw',                      
            '名称': 'name',
            '子件名称':'name',
            '用量': 'num',
            '数量': 'num',
            '基本用量':'num',            
            '材料': 'metal',
            '单重':'weight',
            '备注': 'remark',
            '材料成本': 'cost_mt',
            '人工成本': 'cost_lb',     
            '费用成本': 'cost_exp', 

        }

        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到BOM属性列(层次，编码，名称，数量)'
                break         
            for c, value in enumerate(row):
                for key in title:
                    if isinstance(value, str) and key == value.replace(' ', ''):
                        if title[key] in col:
                            if title[key] == 'lv':  # 检查属性是否重复
                                col['lv_end'] = c
                            else:                        
                                rst['error']='%s 的属性列重复'% key
                                col.clear()
                                return
                        else:
                            col[title[key]] = c

            if ('code' and 'name') in col:
                col['str_row'] = str_row
                if 'lv' in col and 'lv_end' not in col:
                    col['lv_end'] = col['lv']
                break
            else:
                col.clear()
    
    def get_lv(lvs):
        if not excel_bom:   #先要找到ROOT才能继续
            if isinstance(lvs[0], str) and lvs[0].upper() == 'ROOT':                
                return 1
            else:                
                return ' 行没有ROOT'
        elif len(lvs)==1:    #针对层次只有1列：一种是'+++'，另一种是本身只有一层1，2，3，4
            if isinstance(lvs[0], str):                
                lv = len(lvs[0])
            elif isinstance(lvs[0], int):
                lv = 2
            else:
                lv = ' 缺少层次'
            return lv
        else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
            lv_1 = []
            for col, n in enumerate(lvs):
                if isinstance(n, int):
                    lv = col+1
                    lv_1.append(lv)
            if len(lv_1) == 1:
                return lv
            elif len(lv_1) > 1:
                return ' 行层次重复'
            else:
                return ' 行缺少层次'

    def fmt(x):
        if isinstance(x, (int,float)):
            return round(x, 2)        
        elif isinstance(x, datetime):
            return x.strftime('%Y-%m-%d')
        elif isinstance(x, str):
            x = x.replace(' ', '')
            if x:
                try:
                    n = float(x)
                    return round(n, 2)
                except:
                    return x.upper()
            else:
                return '-'
        else:
            return '-'    

    def read_item(wsheet):
         
        row_num = col['str_row'] - 1
        chd_lv=0
        for row in wsheet.iter_rows(min_row=col['str_row'], values_only=True):
            item=[]
            row_num += 1
            for key in title:                
                if key == 'lv' and 'lv' in col:
                    lv = get_lv(row[col['lv']:col['lv_end'] + 1])                         
                else:
                    x=fmt(row[col[key]]) if key in col else '-'
                    item.append(x)

            if 'lv' in col:
                if isinstance(lv, int) and isinstance(item[3], (int, float)):
                    if len(excel_bom)==1:   # 第二行的层次必须是2, 所以要根据第二行的层次来确定一个整体层次的调整系数
                        chd_lv = 2 - lv
                    item.insert(0,lv+chd_lv)
                    
                elif item[2] != '-':
                    if not isinstance(item[3], (int, float)):
                        rst['error']='第 ' + str(row_num) + '没有数量或格式不对'                
                    elif not isinstance(lv, int):
                        rst['error'] = '第 ' + str(row_num) + lv 
                    break
            
            if item[3] != '-':
                item.append(row_num)
                excel_bom.append(item)            

    def excel_bom_total():
        # 计算反查物料在顶层的总用量,和本层用量 [0层次,1编码,2图号,3名称,4数量,5材料,6备注]
        lv_num = {0:1,1: 1} 
        for item in excel_bom:
            lv_num[item[0]] = item[4] * lv_num[item[0] - 1]
            item.insert(5, lv_num[item[0]])

    excel_bom = []
    col = {}    
    rst={}
    skip_sheet = ''
    
    title = ['lv', 'code', 'draw', 'name', 'num', 'metal','weight', 'remark', 'cost_mt', 'cost_lb', 'cost_exp']
    wbook = xl.load_workbook(file, read_only=True)
    
    if sheet == 'ALL':
        names=wbook.sheetnames        
    else:
        names = [wbook.active.title]
           
    for sname in names:        
        wsheet = wbook[sname]
        get_col(wsheet)
        if col: 
            read_item(wsheet)
        else:
            skip_sheet += ws.title
            rst['error'] = '找不到必要的属性列'
    
        
    if 'error' in rst:
        rst['error'] = wsheet.title + ' 表的' + rst['error']
    else:
        head=[]
        for key in title:
            if key in col:
                head.append(key)
            else:
                head.append('-')     
        if 'lv' in col:
            excel_bom_total()
            head.insert(5, 'tot_num')
        
        excel_bom.insert(0,head)        
        rst['bom'] = excel_bom
        
    return rst  #[0层次,1编码,2图号,3名称,4数量,5总数量,6材料,7备注] 


path=r'D:\work\python\excel处理\excel\GPF-33-EBOM.xlsx'
bom = read_design_BOM(path)

for item in bom['bom']:
    print(item)
