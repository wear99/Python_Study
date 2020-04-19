# 3.21 V1.0完成带子层结构BOM的查询功能
# 3.22 添加读取小批BOM功能,然后写入小批BOM库和原始结构库
# 3.26 添加物料读取,并写入物料库
# 3.27 增加在窗口列表查询功能
# 3.29 读取设计BOM,并匹配编码,在设计BOM中查找物料
# 4.5 增加成本读取和匹配功能
# 4.9 改为统一函数读取excel,并按统一格式输出

import tkinter as tk
from tkinter import ttk
import tkinter.filedialog
import tkinter.messagebox
from tkinter.simpledialog import askstring, askinteger, askfloat
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import json
import re
from datetime import datetime
# -----------------弹出窗体GUI程序---------------
class POP_readcode():
    def __init__(self, parent,root):
        self.pop = tk.Toplevel(root)
        self.parent = parent
        
        self.pop.title('更新物料库')
        self.pop.geometry('500x400')
        self.pop.transient(root)
        self.pop.grab_set()  # 聚焦在此窗口上，其它窗口不可用
        self.setpag()
        self.pop.update()   #立即更新窗口，否则会等到程序全部执行完才会更新
        
        self.pop_action()     

    def setpag(self,):
        fm1 = ttk.Frame(self.pop,height = 25)
        fm2 = ttk.Frame(self.pop)
        fm3 = ttk.Frame(self.pop,height = 25)

        fm1.pack()        
        fm3.pack()
        fm2.pack(expand='yes', fill='x')

        self.lab_pop_tit= tk.StringVar()
        ttk.Label(fm1, textvariable=self.lab_pop_tit,font=("微软雅黑", 20,'bold','italic')).pack(pady=10)

        self.lab_pop_name = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_pop_name).pack(padx=20, pady=5, side='left')

        self.lab_pop_rst = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_pop_rst).pack(padx=20, pady=5, side='left')
        
        self.lab_pop_update =tk.StringVar()
        ttk.Label(fm3, textvariable=self.lab_pop_update).pack(padx=20, pady=5,)

    def pop_action(self,):
        def pop_quit(event):
            self.pop.destroy()

        self.lab_pop_tit.set('正在读取...')
        read_code = self.read_code_GUI()
        if read_code:
            self.lab_pop_tit.set('正在更新数据文件...')
            self.pop.update()
            new_code = self.update_code_GUI(read_code)
            if new_code:
                self.parent.new_code = new_code

            self.lab_pop_tit.set('更新完成，按任意键返回...')
        else:
            self.lab_pop_tit.set('读取文件失败，按任意键返回...')

        self.pop.bind('<Any-KeyPress>',pop_quit)
        
    def read_code_GUI(self,):
        def file_list():
            #path = '\\\Sstech\\erp info\\Code\\2010-12-13开始使用新编码111\\'
            path='D:\\work\\python\\excel处理\\excel\\'
            filename = [
                'OEM&集成系统&能效系统加工件新编码.xlsx',        
                '槽烫加工件新编码.xlsx',
                '干洗加工件新编码.xlsx',
                '干衣机加工件新编码.xlsx',        
                '滚筒烫平机加工件新编码.xlsx',
                '水洗加工件新编码.xlsx',
                '折叠机加工件新编码.xlsx',
                '备品备件新编码.xlsx.xlsx',
                '标贴和铭牌新编码.xlsx',
                '各种采购件新编码.xlsx',
                '原材料新编码.xlsx',
                ]
            files = [path + x for x in filename]

            return files
       
        files = file_list()
        read_code={}
        msg = '物料库:\n\n'
        msg1='状态\n\n'
        self.lab_pop_name.set(msg)
        
        for file in files:
            name=file.split('\\')[-1]
            msg+=name + ':\n'
            self.lab_pop_name.set(msg)
            self.pop.update()

            rst=read_excel_code(file)            
            if 'load' in rst:
                msg1 += '读取失败;\n'
                
            elif 'code' in rst:
                msg1 += '成功：' + str(len(rst['code'])) + '条;'
                read_code.update(rst['code'])
                if 'skip' in rst:
                    msg1 += '\t 跳过工作表: ' + rst['skip'] + '\n'
                else:
                    msg1+='\n'
                
            self.lab_pop_rst.set(msg1)
            self.pop.update()

        return read_code
    
    def update_code_GUI(self, read_code):
            rst_code=[]
            rst = update_to_code_db(read_code)
            msg=''
            if 'error' in rst:
                msg='数据文件写入失败'
            else:                
                if 'new' in rst:
                    msg='新增的物料：%d\n' % (len(rst['new']))
                    for key in rst['new']:
                        key.insert(0, 1)
                        key.append('新增')
                        rst_code.append(key)                    
                if 'mod' in rst:
                    msg += '修改的物料：%d' % len(rst['mod'])
                    for key in rst['mod']:
                        key.insert(0, 1)
                        key.append('修改')
                        rst_code.append(key)                    
                else:
                    msg='物料库完成相同，未进行更新...'
                self.lab_pop_update.set(msg)
                return rst_code

# -----------------主窗体GUI程序---------------
class main_GUI():   # 把整个GUI程序 封装在一个类里面
    def __init__(self,):    # 窗体定义，基本函数，其它的都靠它来触发
        self.root = tk.Tk()
        self.root.title('物料查询')
        self.root.geometry('800x600')
        self.setpag()
        self.menu_en1()
        self.menu_tree()
        self.menu_bar()
        #ttk.Style().theme_use('clam')   #('clam','alt','default','classic')
        ttk.Style().configure("Treeview", background="#383838", 
                fieldbackground="black", foreground="white")
        
    def setpag(self,):    # 把界面内容放在一个一起了，便于修改
        fm0 = ttk.Frame(self.root)
        fm1 = ttk.Frame(self.root)
        fm2 = ttk.Frame(self.root)
        fm0.pack()
        fm1.pack()
        fm2.pack(padx=10, expand='yes', fill='both')

        self.eny_t = tk.StringVar()
        #self.target=tk.StringVar()
        self.en1 = ttk.Entry(fm0, width=30, textvariable=self.eny_t)
        self.en1.pack(padx=20, pady=10, side='left')

        self.en1.bind('<Button-3>', self.R_click_en1)
        self.en1.bind("<Return>", self.en1_enter)
        #对于和事件绑定的函数,会自动给个event参数,所有在定义时要加上event参数
        ttk.Button(fm0,
                   text='设计物料查询',command=lambda:self.find_design_code_GUI(self.en1.get())
                   ).pack(
                       padx=20, pady=10, side='right')

        ttk.Button(fm0,
                   text='小批物料查询',
                   command=self.en1_enter).pack(
                       padx=20, pady=10, side='right')                    
                       
        self.lab_r = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_r,font=("微软雅黑", 12,'italic')).pack(pady=5)
        
        self.tev = ttk.Treeview(fm2, columns=('1', '2', '3', '4', '5','6','7'))        
        
        self.tev.heading('#0', text='层次/序号')
        self.tev.heading('1', text='编码')
        self.tev.heading('2', text='图号')
        self.tev.heading('3', text='名称')
        
        self.tev.column('#0', width=120, anchor='w', stretch='no')
        self.tev.column('1', width=100, anchor='w', stretch='no')
        self.tev.column('2', width=100, anchor='w', stretch='no')
        self.tev.column('3', width=300, anchor='w',stretch='no')        

        self.vbar = ttk.Scrollbar(fm2,
                                  orient='vertical',
                                  command=self.tev.yview)
        self.tev.configure(yscrollcommand=self.vbar.set)
        self.vbar.pack(side='right', fill='y')

        self.hbar = ttk.Scrollbar(
            fm2, orient='horizontal', command=self.tev.xview)
        self.tev.configure(xscrollcommand=self.hbar.set)
        self.hbar.pack(side='bottom', fill='x')

        self.tev.pack(expand='yes', fill='both')
        self.tev.bind('<Button-3>', self.R_click_tree)
#-----------------以下窗口动作触发------------------------------
    def menu_en1(self,):
        def onpaste(event=None):
            self.en1.event_generate('<<Paste>>')

        def copy(event=None):
            self.en1.event_generate("<<Copy>>")

        def cut(event=None):
            self.en1.event_generate("<<Cut>>")

        self.menu = tk.Menu(self.root, tearoff=0)
        self.menu.add_command(label="剪切", command=cut)
        self.menu.add_separator()
        self.menu.add_command(label="复制", command=copy)
        self.menu.add_separator()
        self.menu.add_command(label="粘贴", command=onpaste)

    def menu_tree(self,):    # 定义了treeview处的右键菜单内容，但菜单弹出要由post来调用
        def tree_copy(x):
            self.root.clipboard_clear()
            self.root.clipboard_append(x)

        self.menu1 = tk.Menu(self.root, tearoff=0)
        self.menu1.add_command(
            label="复制编码", command=lambda: tree_copy(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="复制图号", command=lambda: tree_copy(self.tree_draw))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="反查BOM", command=lambda: self.find_father_GUI(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="查询子零件", command=lambda: self.find_child_GUI(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(label="在当前BOM中查询", command=self.find_treebom_GUI)
        self.menu1.add_separator()        
        self.menu1.add_command(label="导出列表",command=self.tree_save)
        
    def menu_bar(self,):   # 定义菜单栏   
        m_bar = tk.Menu(self.root)  # 创建菜单组

        m_file = tk.Menu(m_bar, tearoff=0)  # 创建2级菜单组
        m_file.add_separator()
        m_file.add_command(label='导入ERP BOM',command=self.read_batch_bom_GUI)
        m_file.add_separator()        
        m_file.add_command(label='导入设计BOM',command=self.read_design_bom_GUI)
        m_file.add_separator()
        m_file.add_command(label='更新物料库',command=self.read_code_GUI)
        m_file.add_separator()
        m_file.add_command(label='导入成本文件',command=self.read_cost_GUI)
        m_file.add_separator()       
        # mabr上添加一个标签,链接到file_m
        m_bar.add_cascade(label='读取EXCEL文件', menu=m_file)

        m_cost = tk.Menu(m_bar, tearoff=0)
        m_cost.add_separator()
        m_cost.add_command(label='导入成本文件', command=self.read_cost_GUI)
        m_cost.add_separator()
        m_cost.add_command(label='更新组合件成本', command=update_parent_cost)
        m_cost.add_separator()
        m_cost.add_command(label='查看当前物料成本', command=self.tree_add_cost)

        m_bar.add_cascade(label='成本', menu=m_cost)

        m_view = tk.Menu(m_bar, tearoff=0)  # 创建2级菜单组
        root_b = tk.StringVar()
        batch_rootlable=all_batch_bom['index']
        for root,name in batch_rootlable.items():
            # 单选菜单整组有一个variable，每个选项都有一个value。当被选中时，该菜单的value就会赋值给variable。
            m_view.add_radiobutton(label=root +' '+ name, value=root, variable=root_b,
                                   indicatoron=False, command=lambda: self.find_child_GUI(root_b.get()))
            m_view.add_separator()
        m_bar.add_cascade(label=' 查看已导入的小批BOM ', menu=m_view)

        m_view_d = tk.Menu(m_bar, tearoff=0)  # 创建2级菜单组
        root_d = tk.StringVar()        
        for root in all_design_bom:
            # 单选菜单整组有一个variable，每个选项都有一个value。当被选中时，该菜单的value就会赋值给variable。
            m_view_d.add_radiobutton(label=root +' '+ all_design_bom[root][0][3], value=root, variable=root_d,
                                   indicatoron=False, command=lambda: self.view_designbom_GUI(root_d.get()))
            m_view_d.add_separator()        
        m_bar.add_cascade(label=' 查看已导入的设计BOM ', menu=m_view_d)

        m_tool = tk.Menu(m_bar, tearoff=0)
        m_tool.add_separator()
        m_tool.add_command(label='更新Excel编码', command=self.check_excel_GUI)        
        
        m_bar.add_cascade(label='工具', menu=m_tool)

        self.root.config(menu=m_bar)  # 把mbar菜单组 配置到窗体;
      
    def R_click_en1(self, event):
        self.menu.post(event.x_root, event.y_root)   # 在事件坐标处,弹出对应的菜单

    def R_click_tree(self, event):   # 鼠标右键触发程序，该程序通过前面的bind 和右键绑定在一起
        iid = self.tev.identify_row(event.y)   # 返回事件发生时鼠标坐标对应的行
        if iid:   # 如果鼠标所在是空,则不执行右键动作
            self.tev.selection_set(iid)    # 当右键时选中目前鼠标所在的行
            self.tree_code = self.tev.item(self.tev.selection(), 'values')[0]
            self.tree_draw = self.tev.item(self.tev.selection(), 'values')[1]
            self.tree_code = self.tree_code.replace(' old', '')
            self.menu1.post(event.x_root, event.y_root)

    def en1_enter(self, event=None):  #和事件绑定的函数,在事件触发时,会自动给一个event参数,所有定义时必须加上
        x = self.en1.get()
        if x in ('', ' ',None):
            pass
        elif len(x) < 2:
            self.lab_r.set('符合条件物料太多，请补充信息')            
        else:
            self.find_code_GUI(x)

#-----------------以下窗口动作函数--------------------------------        
    def find_code_GUI(self, x):
        x = x.replace('\n', '')  #去除换行符        
        rst_code=find_code(x)
        if len(rst_code[0]) == 1:
            self.lab_r.set(rst_code[0][0])              
        else:
            self.tev.heading('4', text='日期')
            self.lab_r.set(str(x)+' 的物料查询结果:')
            self.tree_out(rst_code,type='CODE')

    def find_father_GUI(self, x):        
        rst_father = find_father_bom(x)
        if len(rst_father[0]) == 1:
            self.lab_r.set(rst_father[0][0])
        else:
            self.tev.heading('4', text='用量')
            self.lab_r.set(str(x)+' 的反查结果:')
            self.tree_out(rst_father,tar=x)

    def find_child_GUI(self, x):        
        rst_child=find_child_bom(x)
        if len(rst_child[0]) == 1:
            self.lab_r.set(rst_child[0][0])
        else:
            self.tev.heading('4', text='用量')
            self.lab_r.set(str(x)+' 的子项结构查询结果')
            self.tree_out(rst_child)        

    def find_design_code_GUI(self,x):
        
        if x in ('', ' ',None):
            pass
        elif len(x) < 2:
            self.lab_r.set('符合条件物料太多，请补充信息')            
        else: 
            rst_d_code = find_design_code(x)
            if len(rst_d_code[0]) == 1:
                self.lab_r.set(rst_d_code[0][0])
            else:
                self.tev.heading('4', text='用量')
                self.lab_r.set(str(x)+' 在设计BOM中查询结果:')
                self.tree_out(rst_d_code,tar=x)
            
    def read_batch_bom_GUI(self,):        
        file_name = tk.filedialog.askopenfilename(title='打开BOM文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = read_design_BOM(file_name)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'bom' in rst:
            root=rst['bom'][0][1]
            if root in all_batch_bom['index']:
                wr1 = tk.messagebox.askquestion(message='该产品代码已存在, 将对当前BOM进行更新\n 是否继续导入?')
                if wr1=='yes':    
                    rst_t1 = update_to_batchbom_db(rst['bom'])
            else:
                rst_t1 = update_to_batchbom_db(rst['bom'])

            if root in all_original_bom:
                wr2 = tk.messagebox.askquestion(message='是否对原始结构库进行覆盖?')
                if wr2=='yes':   
                    rst_t2 = update_to_oribom_db(rst['bom'])
            else:
                rst_t2 = update_to_oribom_db(rst['bom'])

            self.lab_r.set(rst_t1 + ' ; ' + rst_t2)
            self.tree_out(rst['bom'])
            
    def read_design_bom_GUI(self,):
        file_name = tk.filedialog.askopenfilename(title='打开BOM文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = read_design_BOM(file_name)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'bom' in rst:
            #self.tree_out(rst['bom'])
            #wr1 = tk.messagebox.askquestion(message='是否进行编码匹配?')
            #if wr1=='yes':   
            check_code(rst['bom'])
            
            self.lab_r.set('编码已匹配完成')
            wr3 = tk.messagebox.askquestion(message='是否导入设计BOM库?')
            if wr3=='yes':    
                root=rst['bom'][0][1]
                if root in all_design_bom:
                    wr2 = tk.messagebox.askquestion(message='产品代码已存在,是否进行覆盖?')
                    if wr2=='yes':   
                        rst_t1 = update_to_designbom_db(rst['bom'])
                    else:
                        rst_t1='已取消写入设计BOM库'
                else:
                    rst_t1 = update_to_designbom_db(rst['bom'])
            else:
                rst_t1='已取消写入设计BOM库'
            self.lab_r.set(rst_t1)
            self.tree_out(rst['bom'])

    def read_cost_GUI(self,):
        file_name = tk.filedialog.askopenfilename(title='打开成本文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = read_cost(file_name)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'read' in rst:
            wr = tk.messagebox.askquestion(message='是否导入成本库?')
            if wr=='yes':    
                rst1 = update_to_cost_db(rst['read'])
                if 'error' in rst1:
                    self.lab_r.set(rst1['error'])
                elif 'change' in rst1:
                    self.lab_r.set('已成功导入,成本变化的编码:')
                    self.tree_out(rst1['change'])
                else:
                    self.lab_r.set('已成功导入:')
                    #rst['read'].insert(0,[1,'编码','名称','材料成本','人工成本','管理成本','总成本'])
                    self.tree_out(rst['read'])
            else:
                self.lab_r.set('读取的成本数据,未导入')
                self.tree_out(rst['read'])

    def view_designbom_GUI(self, root):
        self.lab_r.set(str(root)+' 的设计BOM')
        read_bom = all_design_bom[root]
        self.tree_out(read_bom)

    def read_code_GUI(self,):
        self.new_code = []
        popcode = POP_readcode(self, self.root)

        self.root.wait_window(popcode.pop)

        if self.new_code:
            self.tree_out(self.new_code)

    def check_excel_GUI(self,):
        file_name = tk.filedialog.askopenfilename(title='打开BOM文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = check_excel_code(file_name)
        if 'error' in rst:
            self.lab_r.set(rst['error'])

    def find_treebom_GUI(self,):
        f = askstring("在列表中查找", "请输入要查询的内容")
        f=f.upper().strip()  # 转大写，去收尾空格
        f = f.replace('\n', '')  #去掉换行符
        
        if f:
            self.tree_out(self.tree_bom,tar=f)          

    def tree_out(self, bom,type='BOM',tar=''):  # 向treeview中写入列表内容
        # BOM格式[0层次,1编码,2图号,3名称,4数量,5本层数量]
        # 设计BOM格式[0层次,1编码,2图号,3名称,4数量,5本层数量,6材料,7备注]
        # 物料格式[0层次,1编码,2图号,3名称,4日期]
        # 成本格式[0层次,1编码,2图号,3名称,4材料成本,5人工成本,6管理成本,7总成本]
        #self.lab_r.set(self.target)
        def set_tree(type):
            if 'CODE' in type:                
                self.tev.heading('4', text='时间')                
                self.tev.column('4', width=100, anchor='w',stretch='no')
                if 'COST' in type:
                    self.tev.heading('4', text='合计成本')
                    self.tev.heading('5', text='材料成本')
                    self.tev.heading('6', text='人工成本')
                    self.tev.heading('7', text='管理费用')
                    self.tev.column('4', width=60, anchor='w', stretch='no')
                    self.tev.column('5', width=60, anchor='w', stretch='no')
                    self.tev.column('6', width=60, anchor='w', stretch='no')
                    self.tev.column('7', width=60, anchor='w',stretch='no')
                
            elif 'BOM' in type:
                self.tev.heading('4', text='数量')
                self.tev.heading('5', text='总数量')
                self.tev.heading('6', text='材料')
                self.tev.heading('7', text='备注')
                self.tev.column('4', width=50, anchor='w', stretch='no')
                self.tev.column('5', width=50, anchor='w', stretch='no')
                self.tev.column('6', width=100, anchor='w', stretch='no')
                self.tev.column('7', width=80, anchor='w', stretch='no')
                
                if 'COST' in type:
                    self.tev.heading('7', text='合计成本')                   
                    self.tev.heading('6', text='单件成本')

        def set_tags(item1):
            for s in self.tev.item(item1, 'values'):
                if tar in s:
                    self.tev.item(item1, tag='tar')
                    rst.append(self.tev.item(item1, 'values'))

                    p1 = self.tev.parent(item1)
                    p2 = self.tev.parent(p1)
                    p3 = self.tev.parent(p2)
                    
                    self.tev.item(p1, open=True)
                    self.tev.item(p2, open=True)
                    self.tev.item(p3, open=True)
                    
                    break

        set_tree(type)
        self.tree_bom = bom
        self.tree_type=type     
        for item in self.tev.get_children():  # 对treeview进行清空
            self.tev.delete(item)
        rst=[]
        lv = {0: ''}
        a = bom[0][0] - 1
        lv_n = [0,0,0,0,0,0,0,0,0]        

        for key in bom:            
            lv_n[key[0] + 1:8] = 0,0,0,0,0,0,0,0
            lv_n[key[0]] += 1

            i = key[0] - a
            lv[i] = self.tev.insert(lv[i - 1],'end',text='+' * key[0] + ' '+str(lv_n[key[0]]),
                                        values=key[1:])
            if tar:
                set_tags(lv[i])

        self.tev.tag_configure('tar', foreground='blue',background='#eeeeff', font='微软雅黑')

        if rst:
                t1='%s在列表中共出现: %d 次'%(tar,len(rst))
                if len(rst[0])>=5:
                    sum = 0
                    try:
                        for x in rst:
                            sum += float(x[4])
                        t1 += ',总用量为: %d' % sum
                    except:
                        pass
                    self.lab_r.set(t1) 
        elif tar:
            t1='%s在列表中未找到'% tar
            self.lab_r.set(t1)                              
            
    def tree_add_cost(self,):   # 在当前显示的物料后添加成本数据
        cost_bom=[]
        if self.tree_type=='BOM':
            for item in self.tree_bom:
                if item[1] in all_cost:
                    tot = all_cost[item[1]][3]
                    item = item[:6] + [round(tot,2), round(tot * item[5],2)]
                else:
                    item = item[:6] + ['-','-']
                cost_bom.append(item)
            self.tree_out(cost_bom,'BOM-COST')
        elif self.tree_type == 'CODE':
            for item in self.tree_bom:
                if item[1] in all_cost:
                    cm = all_cost[item[1]][0]
                    cl = all_cost[item[1]][1]
                    ce = all_cost[item[1]][2]
                    ct = all_cost[item[1]][3]
                else:
                    cm=ct=cl=ce='-'
                item = item[:4] + [ct,cm,cl,ce]
                cost_bom.append(item)
            self.tree_out(cost_bom,'CODE-COST')            

    def tree_save(self,):
        file = tk.filedialog.asksaveasfilename(defaultextension=".xlsx",title='保存文件',
                                                  filetypes=[('xlsx', '*.xlsx')])
        info = save_to_excel(file, self.tree_bom, self.lab_r.get())
        if info == 'ok':
            msg = '已成功导出到文件：' + file
        elif info=='false':
            msg = '写入文件失败'
        else:
            msg = '查询结果导出失败'
        self.lab_r.set(msg)

#-----------------以下主程序函数------------------------------ 
def find_father_bom(f):    # 根据编码反查使用的BOM
    # 在字典每个值的里面查找编码,找到后将对应的key,再作为编码进行同样查找,直到key=index
    
    def find_in_bom(x, n):
        for key, item in all_batch_bom.items():
            if key == 'index':
                continue
            for code in item:
                if x in code:
                    rst_1[n] = [n] + code
                    if key in all_batch_bom['index']:
                        if key not in father_bom_dict:
                            father_bom_dict[key] = []
                        rst_1[n + 1] = [n + 1, key, 1]
                        num = n + 2
                        rst_2 = []
                        for y in rst_1[n + 1:0:-1]:
                            rst_2.append([num - y[0], y[1], y[2]])
                        father_bom_dict[key].append(rst_2[:])
                        break
                    else:
                        find_in_bom(key, n + 1)
                        break

    def fmt_father_bom():    # 去除相同的父项,并由字典转化为列表形式
        for key in father_bom_dict:
            lvcode = {}
            father_bom_dict[key].sort()
            for items in father_bom_dict[key]:
                for item in items:
                    if lvcode.get(item[0], 'NA') != item[1]:
                        lvcode[item[0]] = item[1]
                        for n in range(item[0] + 1, 7):
                            lvcode[n] = ''
                        rst_father_bom.append([item[0], item[1], item[2]])

    def father_bom_total(x):
        # 计算反查物料在顶层的总用量,和本层用量
        lv_num = {0: 1, 1: 1}
        code_num = [0]
        root_index = 0

        for n, item in enumerate(rst_father_bom):
            lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
            item.append(lv_num[item[0]])

            if item[1] == x:
                code_num.append(lv_num[item[0]])

            if item[0] == 1 or n == len(rst_father_bom) - 1:
                rst_father_bom[root_index][3] = (sum(code_num))
                code_num = [0]
                lv_num = {0: 1, 1: 1}
                root_index = n

    rst_father_bom=[]    
    father_bom_dict = {}
    rst_1 = ['' for x in range(7)]

    find_in_bom(f, 1)

    if father_bom_dict:
        fmt_father_bom()
        father_bom_total(f)

        for item in rst_father_bom:
            draw,name=get_code_info(item[1])
            item.insert(2, draw)
            item.insert(3,name)
    
        return rst_father_bom
    else:
        return [['此物料没有在已导入的BOM中使用']]

def find_code(f,db='all',item='all',exact=False):  # 根据输入内容查找物料，db选择物料数据库；exact为TRUE则准确匹配，false则模糊查找；
    rst_code_1 = []
    
    if item == 'code':
        st = 0
        ed = 1
    elif item == 'draw':
        st = 1
        ed = 2
    elif item == 'name':
        st = 2
        ed = 3
    else:
        st = 0
        ed = 4

    if db=='all':
        code_db = all_code
          
    f=f.upper().strip()  # 转大写，去收尾空格
    f=f.replace('\n','')  #去掉换行符
    if f in all_code:
        rst_code_1.append(all_code[f][:])

    elif exact:  # 准确匹配，要求完全相同
        if ed - st == 1:
            for item in list(code_db.values()):
                if f in item[st]:
                    rst_code_1.append(item[:])
        else:    
            for item in list(code_db.values()):
                if f in item[st:ed]:
                    rst_code_1.append(item[:])
    else:
        f = f.replace('.', '\.')     # 点保持原样，不要被改成通配符        
        f = f.replace('*', '.*')    # 将windows习惯用法的 * 转换为python中的 .*
        f = f.replace(' ','.*')       # 字符中间空格，改为通配符
        x = re.compile(f)            # 使用正则表达式中通配符进行查询
        for item in list(code_db.values()):
            for m in item[st:ed]:  # [code,draw,name,date]                
                if m!='-' and x.search(m):
                    rst_code_1.append(item[:])   #要对元素进行添加,而不是整个地址引用,那样会造成原列表被修改
    
    if rst_code_1:  #添加层次,便于输出到GUI
        rst_code_1.sort(key=lambda x: x[0])
        for key in rst_code_1:
            key.insert(0, 1)
        return rst_code_1           #[层次,编码,图号,名称,日期]
    else:
        return [['物料库中不存在']]

def find_child_bom(f):      # 根据编码查找子零件
    def child_bom_total():
        # 计算反查物料在顶层的总用量,和本层用量
        lv_num = {0: 1, 1: 1}       
        for item in rst_child_bom:
            lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
            item.append(lv_num[item[0]])

    rst_child_bom=[]
    def find_child_code(x, n):
        if x in all_batch_bom:
            for item in all_batch_bom[x]:
                rst_child_bom.append([n]+item[:])
                find_child_code(item[0],n+1)

    if f in all_batch_bom:
        rst_child_bom.append([1,f,1])
        find_child_code(f, 2)
        child_bom_total()

        for item in rst_child_bom:
            draw,name=get_code_info(item[1])
            item.insert(2, draw)
            item.insert(3,name)
    
        return rst_child_bom
    else:        
        return [['BOM库中无此物料或没有子零件']]

def find_design_code(f):    #在设计BOM中查找物料
    rst_code_1 = []
    f=f.replace('.','\.')    
    f = f.replace('*', '.*')    # 将windows习惯用法的 * 转换为python中的 .*
    x = re.compile(f)  # 使用正则表达式中通配符进行查询
    for key in all_design_bom:
        for item in all_design_bom[key]:
            for m in item[1:]:
                if isinstance(m,str) and x.search(m):
                    rst_code_1.append(item[:])   #要对元素进行添加,而不是整个地址引用,那样会造成原列表被修改
    
    if rst_code_1:        
        for item in rst_code_1:
            del item[0]
            item.insert(0, 1)            
            
        rst_code_2 = set(tuple(x) for x in rst_code_1)
        rst_code_2=[list(x) for x in rst_code_2]
        rst_code_2.sort(key=rst_code_1.index)                
        return rst_code_1  #[层次,编码,图号,名称,数量,总数量,材料,备注]
    else:
        return [['设计BOM库中未找到相关物料']]

def get_code_info(s):       # 根据编码返回图号和名称    
    sa = s.replace('P', '')    
    if sa in all_code:
        return all_code[sa][1],all_code[sa][2]              
    else:
        return '-','-'

def read_design_BOM(wbook):  # 读取excel的bom文件,格式为导出样式,层次为+++
    # 先判断属性列和root,再读取并按统一格式生成列表:(任何列都允许不存在,由后续程序进行判断)
    # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    # 第一个是属性头
    def get_col(wsheet):
        col.clear()
        title_need = {
            '级别': 'lv',
            '层次': 'lv',
            '序号':'lv',
            '编码': 'code',
            '用量': 'num',
            '数量': 'num',            
            '名称': 'name',            
        }

        title_ex = {            
            '图号': 'draw',
            '代号': 'draw',            
            '材料': 'metal',
            '备注': 'remark'
        }

        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到BOM属性列(层次，编码，名称，数量)'
                break         
            for c, value in enumerate(row):
                for key in title_need:
                    if isinstance(value, str) and key in value.replace(' ', ''):
                        if title_need[key] in col and title_need[key] != 'lv':  # 检查属性是否重复
                            rst['error']='%s 的属性列重复'% key
                            col.clear()
                            return
                        col[title_need[key]] = c

            if len(col) == 4:                
                for c, value in enumerate(row):
                    for key in title_ex:
                        if isinstance(value, str) and key in value.replace(
                                ' ', ''):
                            col[title_ex[key]] = c
                break

            else:
                col.clear()

    def get_root(wsheet):
        def fmt(x):
            if x:
                return str(x).upper().strip()
            else:
                return '-'

        row_num = 1
        for row in wsheet.values:
            row_num += 1            
            if row_num > 10:
                rst['error']='前10行没有root项'
                break
            for n, item in enumerate(row):
                if isinstance(item, str) and item.upper() == 'ROOT':
                    code = fmt(row[col['code']])
                    if code=='-':
                        rst['error'] = 'root行没有产品代码'                            
                    else:
                        col['str_col'] = n
                        col['str_row']=row_num
                        
                        lv = 1
                        num=1
                        name = fmt(row[col['name']])                       
                        draw = fmt(row[col['draw']]) if 'draw' in col  else '-'
                        metal = fmt(row[col['metal']]) if 'metal' in col else '-'
                        remark = fmt(row[col['remark']]) if 'remark' in col else '-'
                                                         
                        excel_bom.append([lv, code, draw, name, num, metal, remark])
                        return 

    def read_excel_bom_item(wsheet):
        def get_lv(row_lv):
            if len(row_lv)==1:    #针对层次只有1列：一种是'+++'，另一种是本身只有一层1，2，3，4
                if isinstance(row_lv[0], str):                
                    lv = len(row_lv[0])
                elif isinstance(row_lv[0], int):
                    lv = 2
                else:
                    lv = ' 行缺少层次'
                return lv
            else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
                lv_1 = []
                for col, n in enumerate(row_lv):
                    if isinstance(n, int):
                        lv = col+1
                        lv_1.append(lv)
                if len(lv_1) == 1:
                    return lv
                elif len(lv_1) > 1:
                    return ' 行层次重复'
                else:
                    return ' 行缺少层次'

        def fmt(x):
            if x:
                return str(x).upper().strip()
            else:
                return '-'

        chd_lv=9
        row_num = col['str_row']-1
        for row in wsheet.iter_rows(min_row=col['str_row'], values_only=True):
            row_num += 1
            lv = get_lv(row[col['str_col']:col['lv'] + 1])                       

            code = fmt(row[col['code']])            
            name = fmt(row[col['name']])
            num = row[col['num']]

            draw = fmt(row[col['draw']]) if 'draw' in col  else '-'
            metal = fmt(row[col['metal']]) if 'metal' in col else '-'
            remark = fmt(row[col['remark']]) if 'remark' in col else '-'

            if isinstance(lv, int) and isinstance(num, (int, float)) and name != '-':
                if chd_lv == 9:   # 第二行的层次必须是2, 所以要根据第二行的层次来确定一个整体层次的调整系数
                    chd_lv = 2 - lv
                excel_bom.append([lv+chd_lv, code, draw, name, num, metal, remark])
            elif name != '-':
                if not isinstance(num, (int, float)):
                    rst['error']='第 ' + str(row_num) + '没有数量或格式不对'                
                elif not isinstance(lv, int):
                    rst['error'] = '第 ' + str(row_num) + lv 
                break

    def excel_bom_total():
        # 计算反查物料在顶层的总用量,和本层用量 [0层次,1编码,2图号,3名称,4数量,5材料,6备注]
        lv_num = {0:1,1: 1} 
        for item in excel_bom:
            lv_num[item[0]] = item[4] * lv_num[item[0] - 1]
            item.insert(5, lv_num[item[0]])
                        
    wb = xl.load_workbook(wbook,)
    wsheet = wb.active
    #wsheet=wb.copy_worksheet(wb.active)

    excel_bom = []
    col = {}
    rst={}
    get_col(wsheet)

    if col and 'error' not in rst:  # 查找层次,编码,数量 对应的列数,及起始行数
        get_root(wsheet)
        if 'error' not in rst:
            read_excel_bom_item(wsheet)
        if 'error' not in rst: 
            excel_bom_total()
            rst['bom'] = excel_bom 

    if 'error' in rst:
        rst['error']= wsheet.title +' 表的' + rst['error']
    return rst    #[0层次,1编码,2图号,3名称,4数量,5总数量,6材料,7备注] 

def read_design_BOM_new(file,sheet=0):  # 统一读取各种bom文件,输出格式为统一的
    # 先判断属性列和root,再读取并按统一格式生成列表:(除了名字和编码以外的列都允许不存在,由后续程序进行判断)
    # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    # 
    def get_col(wsheet):
        col.clear()
        title = {
            '级别': 'lv',
            '层次': 'lv',
            '序号': 'lv',
            '编码': 'code',
            '子件编码':'code',
            '图号': 'draw',
            '代号': 'draw',                      
            '名称': 'name',
            '子件名称':'name',
            '用量': 'num',
            '数量': 'num',
            '基本用量':'num',            
            '材料': 'metal',
            '单重':'weight',
            '备注': 'remark',
            '材料成本': 'cost_mt',
            '人工成本': 'cost_lb',     
            '费用成本': 'cost_exp', 

        }

        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到BOM属性列(层次，编码，名称，数量)'
                break         
            for c, value in enumerate(row):
                for key in title:
                    if isinstance(value, str) and key == value.replace(' ', ''):
                        if title[key] in col:
                            if title[key] == 'lv':  # 检查属性是否重复
                                col['lv_end'] = c
                            else:                        
                                rst['error']='%s 的属性列重复'% key
                                col.clear()
                                return
                        else:
                            col[title[key]] = c

            if ('code' and 'name') in col:
                col['str_row'] = str_row
                if 'lv' in col and 'lv_end' not in col:
                    col['lv_end'] = col['lv']
                break
            else:
                col.clear()
    
    def get_lv(lvs):
        if not excel_bom:   #先要找到ROOT才能继续
            if isinstance(lvs[0], str) and lvs[0].upper() == 'ROOT':                
                return 1
            else:                
                return ' 行没有ROOT'
        elif len(lvs)==1:    #针对层次只有1列：一种是'+++'，另一种是本身只有一层1，2，3，4
            if isinstance(lvs[0], str):                
                lv = len(lvs[0])
            elif isinstance(lvs[0], int):
                lv = 2
            else:
                lv = ' 缺少层次'
            return lv
        else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
            lv_1 = []
            for col, n in enumerate(lvs):
                if isinstance(n, int):
                    lv = col+1
                    lv_1.append(lv)
            if len(lv_1) == 1:
                return lv
            elif len(lv_1) > 1:
                return ' 行层次重复'
            else:
                return ' 行缺少层次'

    def fmt(x):
        if isinstance(x, (int,float)):
            return round(x, 2)        
        elif isinstance(x, datetime):
            return x.strftime('%Y-%m-%d')
        elif isinstance(x, str):
            x = x.replace(' ', '')
            if x:
                try:
                    n = float(x)
                    return round(n, 2)
                except:
                    return x.upper()
            else:
                return '-'
        else:
            return '-'    

    def read_item(wsheet):
         
        row_num = col['str_row'] - 1
        chd_lv=0
        for row in wsheet.iter_rows(min_row=col['str_row'], values_only=True):
            item=[]
            row_num += 1
            for key in title:                
                if key == 'lv':
                    if 'lv' in col:
                        lv = get_lv(row[col['lv']:col['lv_end'] + 1])
                    else:
                        item.append(1)                       
                else:
                    x=fmt(row[col[key]]) if key in col else '-'
                    item.append(x)

            if 'lv' in col:
                if isinstance(lv, int) and isinstance(item[3], (int, float)):
                    if len(excel_bom)==1:   # 第二行的层次必须是2, 所以要根据第二行的层次来确定一个整体层次的调整系数
                        chd_lv = 2 - lv
                    item.insert(0,lv+chd_lv)
                    
                elif item[2] != '-':
                    if not isinstance(item[3], (int, float)):
                        rst['error']='第 ' + str(row_num) + '没有数量或格式不对'                
                    elif not isinstance(lv, int):
                        rst['error'] = '第 ' + str(row_num) + lv 
                    break
            
            if item[3] != '-':
                item.append(row_num)
                excel_bom.append(item)            

    def excel_bom_total():
        # 计算反查物料在顶层的总用量,和本层用量 [0层次,1编码,2图号,3名称,4数量,5材料,6备注]
        lv_num = {0:1,1: 1} 
        for item in excel_bom:
            if isinstance(item[4],(int,float)):
                lv_num[item[0]] = item[4] * lv_num[item[0] - 1]
                item.insert(5, lv_num[item[0]])
            else:
                item.insert(5, '-')
 
    #wsheet=wb.copy_worksheet(wb.active)

    excel_bom = []
    col = {}    
    rst={}
    skip_sheet=''
    title = ['lv', 'code', 'draw', 'name', 'num', 'metal','weight', 'remark', 'cost_mt', 'cost_lb', 'cost_exp']
    
    wbook = xl.load_workbook(file,read_only=True)
    if sheet == 'ALL':
        names=wbook.sheetnames        
    else:
        names = [wbook.active.title]
           
    for sname in names:        
        wsheet = wbook[sname]
        get_col(wsheet)
        if col: 
            read_item(wsheet)
        else:
            skip_sheet += wsheet.title
            rst['error'] = '找不到必要的属性列'
                    
        
    if 'error' in rst:
        rst['error'] = wsheet.title + ' 表的' + rst['error']
    else:
        #excel_bom_total()

        head=[]
        for key in title:
            if key in col:
                head.append(key)
            else:
                head.append('-')     
        #head.insert(5, 'tot_num')                
        
        excel_bom.insert(0,head)     
        rst['bom'] = excel_bom
                
    if skip_sheet:
        rst['skip'] = skip_sheet
           
    return rst  #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]

def check_code(excel_bom):   # 检查编码
    # 如果有编码,先检查在编码库中是不是old,如果是则查找新的;如果没有编码且有图号,则根据图号去编码库查找
    def get_new(x, im='all', ex=False):
        if x in all_code and 'old' in all_code[x][0]:
            x1=x.replace('P','')
            codes = find_code(x1[:-1],item=im,exact=ex)
            if len(codes[0])>1:
                codes1 = [i[1] for i in codes]
                codes1.sort()
                return codes1
        else:
            return x

    def get_code(x, im='all', ex=True):
        code = find_code(x, item=im, exact=ex)
        if len(code)==1 and len(code[0]) > 1:
            return code[0][1]
        
    for item in excel_bom:
        code = item[1]
        draw = item[2]
        name = item[3]
        
        if code == '-':
            if draw != '-' and ('GB' or 'JB') not in draw:
                codes=get_code(draw, im='draw',ex=True)
                if codes:
                    code= codes
            elif name != '-':
                codes=get_code(draw, im='name',ex=True)
                if codes:
                    code= codes    
        elif isinstance(code,str) and 'R' in code:
            code = get_new(code, 'code', '')

        item.insert(1, code)
        del item[2]

def check_excel_code(file):
    def get_col(wsheet):
        col.clear()
        title_need = {            
            '编码': 'code',
            '图号': 'draw',
            '代号': 'draw',            
            '名称': 'name',            
        }

        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到BOM属性列(层次，编码，名称，数量)'
                break         
            for c, value in enumerate(row):
                for key in title_need:
                    if isinstance(value, str) and key in value.replace(' ', ''):

                        col[title_need[key]] = c

            if len(col) == 3:
                col['str_row']=str_row
                break
            else:
                col.clear()

    def update_excel():
        def get_new(x, im='all', ex=False):
            if x in all_code and 'old' in all_code[x][0]:
                x1=x.replace('P','')
                codes = find_code(x1[:-1],item=im,exact=ex)
                if len(codes[0])>1:
                    codes1 = [i[1] for i in codes]
                    codes1.sort()
                    return codes1
            else:
                return x

        def get_code(x, im='all', ex=True):
            code = find_code(x, item=im, exact=ex)
            if len(code)==1 and len(code[0]) > 1:
                return code[0][1]
        
        def fmt(x):
            if x:
                return str(x).upper().strip()
            else:
                return '-'

        fill_blue = PatternFill('solid',fgColor='EFBF00')
        row_num=col['str_row']-1
        for row in wsheet.iter_rows(min_row=col['str_row'], values_only=True):           
            row_num+=1
            code = fmt(row[col['code']])            
            name = fmt(row[col['name']])
            draw = fmt(row[col['draw']])

            code1='-'
            if code == '-':
                if draw != '-' and ('GB' or 'JB') not in draw:
                    codes=get_code(draw, im='draw')
                    if codes:
                        code1= codes
                elif name != '-':
                    codes=get_code(name, im='name')
                    if codes:
                        code1= codes    
            elif isinstance(code,str) and 'R' in code:
                code1 = get_new(code, im='code')

            if code1!='-' and code1 != code:
                wsheet.cell(row=row_num, column=col['code'] + 1).value = code1                
                wsheet.cell(row=row_num, column=col['code'] + 1).fill=fill_blue   #对修改过的单元格进行填充


    wbook = xl.load_workbook(file)
    wsheet = wbook.active
    wbook.copy_worksheet(wsheet)   #创建一个原工作表的备份
    
    col = {}
    rst={}
    get_col(wsheet)

    if col and 'error' not in rst:  # 查找层次,编码,数量 对应的列数,及起始行数       
        update_excel()
        try:
            wbook.save(file)
        except:
            rst['error']='文件写入失败'

    if 'error' in rst:
        rst['error'] = wsheet.title + ' 表的' + rst['error']
        
    return rst

def update_to_batchbom_db(excel_bom, mode='A'):  # 将已读取的bom写入当前库,删除没有子零件的物料
    #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    def creat_bom_dict():       
        excel_bom_dict[root] = []
        lv_code = {1:root,}
        ex_code_lv = 10
        
        for n in excel_bom[1:]:
            code = n[1]
            lv = n[0]
            num = n[4]
            if lv > ex_code_lv:
                continue
            else:
                excel_bom_dict[lv_code[lv - 1]].append([code, num])

            if code in excel_bom_dict:
                ex_code_lv = lv
            else:
                lv_code[lv] = code
                excel_bom_dict[code] = []
                ex_code_lv = 10        

    def update_all_batch_bom():
        
        # A 模式: 对于已存在的,如果子项为空,则重新写入.
        # W 模式: 对于已存在的,把原来的清空,按新导入重新添加子项
        if mode == 'W':
            #把excel_bom_dict字典的key,value写入all_batch_bom，相同的进行覆盖
            all_batch_bom.update(excel_bom_dict)            
            
        elif mode == 'A':
            for key, item in excel_bom_dict.items():               
                if item and key not in all_batch_bom:
                    all_batch_bom[key] = item

        if root not in all_batch_bom['index']:
            all_batch_bom['index'][root]=rootname                 
 
    root = excel_bom[0][1]
    rootname = excel_bom[0][3]         
    
    excel_bom_dict={}
    creat_bom_dict()
    update_all_batch_bom()

    try:
        with open('all_batch_bom.json', 'w', encoding='utf-8') as f:
            json.dump(all_batch_bom, f, indent=4, ensure_ascii=False)

        return '已成功更新当前BOM库'
    except:
        return '文件写入失败'

def update_to_oribom_db(excel_bom):   # 将已读取的BOM写入 原始结构库
    root = excel_bom[0][1]
    all_original_bom[root] = excel_bom[:]
    try:   
        with open('all_original_bom.json', 'w', encoding='utf-8') as f:
            json.dump(all_original_bom, f, indent=4, ensure_ascii=False)
        
        return root + '已成功写入原始BOM库'
    except:
        return '写入文件失败'

def update_to_designbom_db(excel_bom):  # 将已读取的设计BOM写入设计BOM库
    #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    root = excel_bom[0][1]   
    rootname = excel_bom[0][3]

    all_design_bom[root] = excel_bom[:]
    try:
        with open('all_design_bom.json', 'w', encoding='utf-8') as f:
            json.dump(all_design_bom, f, indent=4, ensure_ascii=False)
    
        return root + rootname + '已成功写入设计BOM库'
    except:
        return '设计BOM库写入失败'   

def read_excel_code(file):      #读取物料excel文件,然后写入物料库,并标记旧编码
    
    def get_col(ws):
        col = {}
        title_need = {
            '新编码': 'code',            
            '描述': 'name',
            '图号': 'draw',
        }

        start_r = 1
        for row in ws.values:
            start_r += 1
            if start_r > 6:
                break         
            for c, value in enumerate(row):
                for key in title_need:
                    if isinstance(value, str) and key == value.replace(' ', ''):
                        if title_need[key] in col and title_need[key] != 'lv':  # 检查属性是否重复
                            rst['error']='%s 的属性列重复'% key
                            col.clear()
                            return
                        col[title_need[key]] = c

            if len(col) == 3:
                col['start_r'] = start_r
                col['date']=col['draw']+1
                return col
            else:
                col = {}

    def std(s):
        if isinstance(s, datetime):
            return s.strftime('%Y-%m-%d')
        elif s:
            return str(s).upper().strip()
        else:
            return '-'

    def add_item(ws):
        for row in ws.iter_rows(min_row=col['start_r'], values_only=True):
            code = std(row[col['code']])            
            name = std(row[col['name']])

            if isinstance(row[col['draw']], datetime):
                date = std(row[col['draw']])
                draw='-'
            else:
                draw = std(row[col['draw']])
                date= std(row[col['date']])               

            if code != '-':
                excel_code[code]=[code, draw, name, date]
    
    excel_code={}   # 本次读取的物料库    
    skip_sheet = ''    
    rst = {}
    
    try:
        wb = xl.load_workbook(file, read_only=True)
    except:        
        rst['load']='error'
        return rst

    for sname in wb.sheetnames:        
        ws = wb[sname]
        col = get_col(ws)
        if col: 
            add_item(ws)
        else:
            skip_sheet += ws.title
    
    if excel_code:        
        rst['code']=excel_code
            
    if skip_sheet:
        rst['skip'] = skip_sheet  

    return rst

def update_to_code_db(excel_code):
    # 查找旧编码,并在编码前加标记 *
    #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    def old_item_check():
        # 取出字典中所有的key,组成list，然后排序进行比较
        keys=list(all_code.keys())
        keys.sort()
        # 物料按编码排序,对任一编码和下一个编码进行比对,如果除了最后一位相同,且比较小,则认为是旧编码
        for x, a in enumerate(keys):
            if 'old' in all_code[a][0]:
                continue
            if a[2] == 'R':
                if a[-1] == 'P':
                    a = a.replace('P', '')
                b = keys[x + 1]
                if b[-1] == 'P':
                    all_code[a][0] += ' old'            
                elif a[:-1] == b[:-1] and len(a)==len(b):
                    all_code[a][0] +=' old'
            if x == len(keys) - 2:
                break

    def update_code_db():        
        for key in excel_code:
            if key not in all_code:
                code_new.append(excel_code[key])
            elif excel_code[key] != all_code[key] and 'old' not in all_code[key][0]:                
                code_mod.append(excel_code[key])

        all_code.update(excel_code)

    rst={}
    code_new = []
    code_mod=[]
    update_code_db()
    old_item_check()

    if code_new or code_mod:
        try:
            with open('all_code.json', 'w', encoding='utf-8') as f:
                json.dump(all_code, f, indent=4, ensure_ascii=False)        
            if code_new:
                rst['new'] = code_new
            if code_mod:
                rst['mod']=code_mod
        except:
            rst['error']='写入文件失败'

    return rst

def read_cost(file):
    def get_col(wsheet):
        col.clear()
        cost_need = {            
            '编码': 'code',
            '材料成本': 'metal',
            '人工成本': 'labor', 
            '子件名称': 'name',
            '费用成本': 'exp',
        }

        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到BOM属性列(层次，编码，名称，数量)'
                break         
            for c, value in enumerate(row):
                for key in cost_need:
                    if isinstance(value, str) and key in value.replace(' ', ''):
                        if cost_need[key] in col and cost_need[key] != 'lv':  # 检查属性是否重复
                            rst['error']='%s 的属性列重复'% key
                            col.clear()
                            return
                        col[cost_need[key]] = c

            if len(col) == 5:
                col['start_r']= str_row             
                break
            else:
                col.clear()

    def std(s):       
        if isinstance(s, (int, float)):
            return round(s,2)            
        else:
            return 0
        
    def add_item(ws):
        for row in ws.iter_rows(min_row=col['start_r'], values_only=True):
            code = row[col['code']]
            code = code.upper().strip() if code else '-'
            name = row[col['name']]            
            metal = std(row[col['metal']])
            labor = std(row[col['labor']])            
            exp = std(row[col['exp']])
            tot = std(metal + labor + exp)
            
            if code != '-':
                excel_cost.append([1,code,name, metal, labor, exp,tot])
    
    excel_cost = []
    col = {}
    rst = {}    
    try:
        wb = xl.load_workbook(file, read_only=True)
    except:        
        rst['error']=file+'文件读取失败'
        return rst
 
    wsheet = wb.active
    get_col(wsheet)

    if col and 'error' not in rst:  # 查找层次,编码,数量 对应的列数,及起始行数        
        add_item(wsheet)
        rst['read']= excel_cost

    return rst

def update_to_cost_db(excel_cost):
    rst = {}
    cost_change=[]
    for item in excel_cost:
        if item[6]:
            if item[1] in all_cost and all_cost[item[1]][3]!=0 and item[6] != all_cost[item[1]][3]:
                cost_change.append(item)
                cost_change.append([2, '', ''] + all_cost[item[1]])            
            all_cost[item[1]] = item[3:]
             
    try:
        with open('all_cost.json', 'w', encoding='utf-8') as f:
            json.dump(all_cost, f, indent=4, ensure_ascii=False)
        
        rst['info']='已成功更新成本库'        
    except:
        rst['error'] = '文件写入失败'
        
    if cost_change:
        rst['change'] = cost_change

    return rst

def update_parent_cost():  # 对BOM库中组合件的成本重新按结构进行累加
    
    def find_child_cost(x, n,c0,c1,c2,c3):
        if x in all_batch_bom and x not in updated_code and (x[:3] == '24R' or x[0]=='C'):
            #只针对24R组合件或成品吗需要对子零件成本累加,其它的无须更新
            all_cost[x]=[0,0,0,0]
            for item in all_batch_bom[x]:
                all_cost[x]=find_child_cost(item[0], item[1],all_cost[x][0],all_cost[x][1],all_cost[x][2],all_cost[x][3])
            updated_code[x] = ''
            print(x, all_cost[x])
            
        if x in all_cost:
            updated_code[x] = ''
            return round(c0+all_cost[x][0] * n,2),round(c1+all_cost[x][1] * n,2),round(c2+all_cost[x][2] * n,2),round(c3+all_cost[x][3] * n,2)           
        else:
            updated_code[x] = ''
            return c0,c1,c2,c3

    updated_code = {}
    for f in all_batch_bom:
        if f[:3]=='24R'or f[0]=='C':
            all_cost[f] = find_child_cost(f, 1, 0, 0, 0, 0)
    
    for key,item in all_cost.items():  #删除成本中的0值
        if item[3] == 0:
           del all_cost[key]

def read_json_date(filename):  # 从现有文件读取数据
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        print(filename+' 数据库读取失败')
        return {}
            
def wr_root_to_allcode():   # 把已读取BOM的产品代码写入物料库
    if all_batch_bom.get('index',''):
        for root,name in all_batch_bom['index'].items():
            all_code[root] = [root, '-', name, '-']  # 把BOM表中的产品名写入code表

def save_to_excel(file, save_bom,target):  #把表格内容保存到excel文件

    wb = xl.Workbook()
    ws = wb.active

    fill_blue = PatternFill('solid',fgColor='00B2EE') #设置填充颜色为 橙色 
    font_title = Font(u'微软雅黑', size=14, bold=True, italic=True)  #设置字体样式
    font_bold = Font(u'微软雅黑', bold=True,)
    
    ws.append([target])     #添加标题    
    ws['A1'].font = font_title

    ws.append(['层次', '序号', '编码', '图号', '名称', '数量', '总数量', '材料', '备注'])
    for i in ws[2]:
        i.fill = fill_blue
        i.font = font_bold
    
    lv_n = [0,0,0,0,0,0,0,0,0]       
    for key in save_bom:
        lv_n[key[0] + 1:8] = 0,0,0,0,0,0,0,0
        lv_n[key[0]] += 1

        key[0]='+'*key[0]
        key=[x if x!='-' else '' for x in key]
        ws.append(['+'*key[0],lv_n[key[0]]]+key[1:])

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20 
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15
    try:
        wb.save(file)
        return 'ok'
    except:
        return 'false'
        
#---------------------主程序区---------------------
all_root = {
    'C07-0024': '卓越CEF-33-1L-FS折叠堆码机',
    'C07-0026': '卓越CEF-33-1L-F-FS折叠机堆码机',
    'C07-0031': 'GPF-33-4L-FS 全自动折叠堆码机',
    'C08-0005': 'CSJ-150穿梭机',
    'C07-0020': '卓越CETF-20毛巾折叠机',
    'C07-0010': '卓越CEF-30-2L折叠机',
    'C07-0019': '卓越CEF-33-4L-FS折叠堆码机'
}

all_code = read_json_date('all_code.json')
all_batch_bom = read_json_date('all_batch_bom.json')
all_design_bom=read_json_date('all_design_bom.json')
all_original_bom = read_json_date('all_original_bom.json')
all_cost = read_json_date('all_cost.json')

if not all_batch_bom:
    all_batch_bom['index'] = {}

wr_root_to_allcode()

print('code库记录: ', len(all_code))
print('成本库记录:',len(all_cost))
print('bom库记录: %d ,已读取的产品有:' % len(all_batch_bom), all_batch_bom['index'])

op = main_GUI()
op.root.mainloop()





