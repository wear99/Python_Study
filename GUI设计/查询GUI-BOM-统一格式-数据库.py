# -*- coding: utf-8 -*-
# 3.21 V1.0完成带子层结构BOM的查询功能
# 3.22 添加读取小批BOM功能,然后写入小批BOM库和原始结构库
# 3.26 添加物料读取,并写入物料库
# 3.27 增加在窗口列表查询功能
# 3.29 读取设计BOM,并匹配编码,在设计BOM中查找物料
# 4.5 增加成本读取和匹配功能
# 4.9 改为统一函数读取excel,并按统一格式输出
# 4.12 存储格式由JSON改为sqlite数据库,启动时读取数据库到各字典中
# 4.15 成本库中增加日期，列出变动成本，根据batchbom结构对组合件成本重算，显示时对没有成本的进行计算
# 4.20 excel根据类型读取对应列，并输出，不再按统一格式
# 4.22 完善check code查找功能，更新编码也统一到读取函数中
# 4.23 将designbom改为按小批物料格式存储，新增designcode库，存储15R，16R，24R，28R及没编码物料
# 4.25 完善数据操作,全部完成后再提交,如果出错则回滚
# 4.26 将读取BOM合并,并可以选择对已有的层次关系是否覆盖还是跳过
# 4.28 增加图纸路径读取保存,物料根据图号打开对应的图纸

import tkinter as tk
from tkinter import ttk
import tkinter.filedialog
import tkinter.messagebox
from tkinter.simpledialog import askstring, askinteger, askfloat
import openpyxl as xl
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
import json, sqlite3, re
from os import path,listdir,startfile

from datetime import datetime
# -----------------弹出窗体GUI程序---------------
class POP_readcode():
    def __init__(self, parent,root):
        self.pop = tk.Toplevel(root)
        self.parent = parent
        
        self.pop.title('更新物料库')
        self.pop.geometry('500x400')
        self.pop.transient(root)
        self.pop.grab_set()  # 聚焦在此窗口上，其它窗口不可用
        self.setpag()
        self.pop.update()   #立即更新窗口，否则会等到程序全部执行完才会更新
        
        self.pop_action()     

    def setpag(self,):
        fm1 = ttk.Frame(self.pop,height = 25)
        fm2 = ttk.Frame(self.pop)
        fm3 = ttk.Frame(self.pop,height = 25)

        fm1.pack()        
        fm3.pack()
        fm2.pack(expand='yes', fill='both')

        self.lab_pop_tit= tk.StringVar()
        ttk.Label(fm1, textvariable=self.lab_pop_tit,font=("微软雅黑", 20,'bold','italic')).pack(pady=10)

        self.lab_pop_name = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_pop_name,justify = 'right',anchor = 'n').pack(padx=20, pady=5, side='left',fill='both')

        self.lab_pop_rst = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_pop_rst,anchor = 'n').pack(padx=20, pady=5, side='left',fill='both')
        
        self.lab_pop_update =tk.StringVar()
        ttk.Label(fm3, textvariable=self.lab_pop_update).pack(padx=20, pady=5,)

    def pop_action(self,):
        def pop_quit(event):
            self.pop.destroy()
        
        read_code = self.read_code_GUI()
        if read_code:
            code_mod = self.update_code_GUI(read_code)
            if code_mod:
                self.parent.code_mod = code_mod               
        
        self.lab_pop_tit.set('按任意键返回...')
        self.pop.bind('<Any-KeyPress>',pop_quit)
        
    def read_code_GUI(self,):
        def file_list():
            path = '\\\Sstech\\erp info\\Code\\2010-12-13开始使用新编码\\'
            #path='D:\work\python\excel处理\excel\'
            filename = [
                'OEM&集成系统&能效系统加工件新编码.xlsx',        
                '槽烫加工件新编码.xlsx',
                '干洗加工件新编码.xlsx',
                '干衣机加工件新编码.xlsx',        
                '滚筒烫平机加工件新编码.xlsx',
                '水洗加工件新编码.xlsx',
                '折叠机加工件新编码.xlsx',
                '备品备件新编码.xlsx',
                '标贴和铭牌新编码.xlsx',
                '各种采购件新编码.xlsx',
                '原材料新编码.xlsx',
                ]
            
            files = [path + x for x in filename]

            return files
        
        files = file_list()
        read_code=[]
        msg = '物料库:\n\n'
        msg1 = '状态\n\n'
        self.lab_pop_name.set(msg)
        self.lab_pop_tit.set('正在读取...')       
        
        for file in files:
            name=file.split('\\')[-1]
            msg+=name + ':\n'
            self.lab_pop_name.set(msg)
            self.pop.update()

            rst=read_design_BOM(file,type='CODE')
            if 'error' in rst:
                msg1 += rst['error']+'\n'
                
            elif 'bom' in rst:
                msg1 += '成功：' + str(len(rst['bom'])) + '条;'
                read_code+=rst['bom']
                if 'skip' in rst:
                    msg1 += '\t 跳过工作表: ' + rst['skip'] + '\n'
                else:
                    msg1+='\n'
                
            self.lab_pop_rst.set(msg1)
            self.pop.update()

        return read_code
    
    def update_code_GUI(self, read_code):
        self.lab_pop_tit.set('正在更新数据文件...')
        self.pop.update()
        rst_code=[]
        rst = update_to_code_db(read_code)
        msg='共读取的物料:%d\n'% len(read_code)
        if 'error' in rst:
            msg=rst['error']
        else:                
            if 'new' in rst:
                msg+='新增的物料：%d\n' % rst['new']                  
            if 'mod' in rst:
                msg += '修改的物料：%d' % (len(rst['mod']) / 2)
                rst_code=rst['mod']
                                
        self.lab_pop_update.set(msg)        
        return rst_code

class edit_root_path():    
    def __init__(self, parent,root):
        self.pop = tk.Toplevel(root)
        self.parent = parent
        
        self.pop.title('修改产品图纸位置')
        self.pop.geometry('500x400')
        self.pop.transient(root)
        self.pop.grab_set()  # 聚焦在此窗口上，其它窗口不可用
        self.setpag()        
        self.menu_tree()
        #self.menu_bar()
        self.get_drawnum()
        self.path_tree_out()        

    def setpag(self,):
        #fm1 = ttk.Frame(self.pop,height = 25)
        fm2 = ttk.Frame(self.pop, height=25)
        fm3 = ttk.Frame(self.pop)
        #fm1.pack()
        fm2.pack()
        fm3.pack(expand='yes', fill='both')
        
        self.path_lab_1 = tk.StringVar()
        ttk.Label(fm2, textvariable=self.path_lab_1,font=("微软雅黑", 12,'italic')).pack(pady=5)
        
        self.tev = ttk.Treeview(fm3, show='headings',columns=('1', '2', '3'),selectmode='browse')
        self.tev.heading('1', text='产品')
        self.tev.heading('2', text='图纸目录')
        self.tev.heading('3', text='图纸数量')
        self.tev.column('#0', width=40, anchor='w')
        self.tev.column('1', width=100, anchor='w')
        self.tev.column('2', width=400, anchor='w')
        self.tev.column('3', width=40, anchor='w')
        
        self.vbar = ttk.Scrollbar(fm3,
                                  orient='vertical',
                                  command=self.tev.yview)
        self.tev.configure(yscrollcommand=self.vbar.set)
        self.vbar.pack(side='right', fill='y')        

        self.tev.pack(expand='yes', fill='both')
        self.tev.bind('<Button-3>', self.right_click)

    def right_click(self, event):
        iid = self.tev.identify_row(event.y)   # 返回事件发生时鼠标坐标对应的行
        # 如果鼠标所在是空,则不执行右键动作
        if iid:
            self.tev.selection_set(iid)    # 当右键时选中目前鼠标所在的行
            self.name = self.tev.item(self.tev.selection(), 'values')[0]
            self.pathdir = self.tev.item(self.tev.selection(), 'values')[1]
            self.menu1.post(event.x_root, event.y_root)
    def menu_bar(self,):   # 定义菜单栏   
        m_bar = tk.Menu(self.pop)  # 创建菜单组
        m_bar.add_command(label="新增目录", command=lambda: self.edit_path(type='ADD'))      

        self.pop.config(menu=m_bar)  # 把mbar菜单组 配置到窗体;
    def menu_tree(self,):  # 定义了treeview处的右键菜单内容，但菜单弹出要由post来调用
        self.menu1 = tk.Menu(self.pop, tearoff=0)
        self.menu1.add_command(
            label="新增目录", command=lambda:self.edit_path(type='ADD'))
        self.menu1.add_separator()
        
        self.menu1.add_command(
            label="修改目录", command=lambda: self.edit_path(type='CHANGE'))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="删除目录", command=lambda: self.edit_path(type='DEL'))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="重新搜索目录",command=lambda:self.edit_path(type='UPDATE'))                
        self.menu1.add_separator()
        self.menu1.add_command(label="导出列表")
    def get_drawnum(self,):
        self.draw_num = {}
        for key in path_root:
            rst = find_db('remark', key, 'drawPATH')
            self.draw_num[key] = len(rst)
            
    def path_tree_out(self,):
        for item in self.tev.get_children():  # 对treeview进行清空
            self.tev.delete(item)
        if path_root:
            for key, item in path_root.items():
                self.tev.insert('', 'end', values=(key, item,self.draw_num[key]))        

    def edit_path(self, type='ADD'):    # 编辑目录,包含新增,更改,删除,更新 4种模式
        def get_name():
            name = askstring('',"请输入产品信息：")
            name = name.upper().strip()  # 转大写，去收尾空格
            name = name.replace('\n', '')  # 去掉换行符
            if name in path_root:
                tk.messagebox.showerror('产品信息已存在，请重新输入！')
            else:
                return name

        def get_path():
            pathdir = tk.filedialog.askdirectory(title='选择图纸文件夹')
            for item in path_root.values():
                if pathdir in item:
                    tk.messagebox.showerror('产品文件夹已存在，请重新选择！')
                else:
                    return pathdir
        rst={}
        type=type.upper()
        if type=='ADD':   # 新增模式下,需输入name,选择目录     
            name = get_name()
            pathdir = get_path()
        elif type == 'CHANGE':  #更改模式:name
            name = self.name
            if name:
                pathdir = get_path()
        elif type == 'UPDATE':
            name = self.name
            pathdir = self.pathdir
        elif type == 'DEL':
            name = self.name
            pathdir = self.pathdir
            
        if not (name and pathdir):
            return

        if type in ('ADD','UPDATE','CHANGE'):   #扫描指定的文件夹
            rst = scan_path(pathdir)
            if 'error' in rst:
                self.path_lab_1.set(rst['error'])
                return
            elif 'path' in rst:
                rst['path']=[tuple(x+[name]) for x in rst['path']]                
                num = len(rst['path'])
                self.draw_num[name] = num
                t1 = name + ' 中读取的图纸数量：' + str(num) + ',已导入数据库'

        if type in ('DEL','UPDATE','CHANGE'):   # 删除保存的文件夹和所有图纸路径
            rst_1 = self.remove_path_db(name)
            if 'error' in rst_1:
                self.path_lab_1.set(rst_1['error'])
                return
            else:
                del path_root[name]
                
        if type in ('ADD','UPDATE','CHANGE') and 'path' in rst:   # 写入新的文件夹和所有图纸路径
            rst = self.update_path_db(name, pathdir, rst['path'])                    
            if 'error' in rst:
                self.path_lab_1.set(rst['error'])
                return
            else:
                path_root[name] = pathdir                    
                self.path_lab_1.set(t1)
        
        if type == 'ADD':
            self.tev.insert('', 'end', values=(name, pathdir, num))
        elif type in ('CHANGE','UPDATE'):
            self.tev.set(self.tev.selection(), column='2', value=pathdir)
            self.tev.set(self.tev.selection(), column='3', value=num)
        elif type == 'DEL':
            self.tev.delete(self.tev.selection())

        #self.path_tree_out()

    def remove_path_db(self, name=''):  # 删除数据库中指定的目录
        rst={}
        old_1 = (name, 'old')
        old_2=('PATH','path')
        try:
            remove_db(old1=old_1, col1='code',old2=old_2,col2='remark',sheet='root')
            remove_db(old1=old_1, col1='remark', sheet='drawPATH')
            conn.commit()
        except Exception as ex:
            conn.rollback()
            rst['error']= '删除时出错：' + str(ex)
        return rst

    def update_path_db(self,root, filedir, file_path):
        rst={} 
        try:
            insert_db(((root,filedir,'PATH'),),'root')
            insert_db(file_path, 'drawPATH')
            conn.commit()
        except Exception as ex:
            conn.rollback()
            rst['error'] = '写入失败: ' + str(ex)
            
        return rst

# -----------------主窗体GUI程序---------------
class main_GUI():   # 把整个GUI程序 封装在一个类里面
    def __init__(self,):    # 窗体定义，基本函数，其它的都靠它来触发
        self.root = tk.Tk()
        self.root.title('物料查询')
        self.root.geometry('800x600')
        self.setpag()
        self.menu_en1()
        self.menu_tree()
        self.menu_bar()
        #ttk.Style().theme_use('clam')   #('clam','alt','default','classic')
        ttk.Style().configure("Treeview", background="#383838", 
                fieldbackground="black", foreground="white")

    def setpag(self,):    # 把界面内容放在一个一起了，便于修改
        fm0 = ttk.Frame(self.root)
        fm1 = ttk.Frame(self.root)
        fm2 = ttk.Frame(self.root)
        fm0.pack()
        fm1.pack()
        fm2.pack(padx=10, expand='yes', fill='both')

        self.eny_t = tk.StringVar()        
        self.en1 = ttk.Entry(fm0, width=30, textvariable=self.eny_t)
        self.en1.pack(padx=20, pady=10, side='left')

        self.en1.bind('<Button-3>', self.R_click_en1)
        self.en1.bind("<Return>", self.en1_enter)
        #对于和事件绑定的函数,会自动给个event参数,所有在定义时要加上event参数
        ttk.Button(fm0,
                   text='设计物料查询',command=lambda:self.find_code_GUI('DESIGN')
                   ).pack(padx=20, pady=10, side='right')

        ttk.Button(fm0,
                   text='小批物料查询',command=self.en1_enter).pack(
                       padx=20, pady=10, side='right')
                       
        self.lab_r = tk.StringVar()
        ttk.Label(fm2, textvariable=self.lab_r,font=("微软雅黑", 12,'italic')).pack(pady=5)
        
        self.tev = ttk.Treeview(fm2, columns=('1', '2', '3', '4', '5','6','7','8'),selectmode='browse')

        self.vbar = ttk.Scrollbar(fm2,
                                  orient='vertical',
                                  command=self.tev.yview)
        self.tev.configure(yscrollcommand=self.vbar.set)
        self.vbar.pack(side='right', fill='y')

        self.hbar = ttk.Scrollbar(
            fm2, orient='horizontal', command=self.tev.xview)
        self.tev.configure(xscrollcommand=self.hbar.set)
        self.hbar.pack(side='bottom', fill='x')

        self.tev.pack(expand='yes', fill='both')
        self.tev.bind('<Button-3>', self.R_click_tree)

    def menu_en1(self,):        # 输入框的右键菜单
        def onpaste(event=None):
            self.en1.event_generate('<<Paste>>')
        def copy(event=None):
            self.en1.event_generate("<<Copy>>")
        def cut(event=None):
            self.en1.event_generate("<<Cut>>")

        self.menu_eny1 = tk.Menu(self.root, tearoff=0)
        self.menu_eny1.add_command(label="剪切", command=cut)
        self.menu_eny1.add_separator()
        self.menu_eny1.add_command(label="复制", command=copy)
        self.menu_eny1.add_separator()
        self.menu_eny1.add_command(label="粘贴", command=onpaste)

    def menu_tree(self,):    # 定义了treeview处的右键菜单内容，但菜单弹出要由post来调用
        def tree_copy(x):
            self.root.clipboard_clear()
            self.root.clipboard_append(x)

        self.menu1 = tk.Menu(self.root, tearoff=0)
        self.menu1.add_command(
            label="复制编码", command=lambda: tree_copy(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="复制图号", command=lambda: tree_copy(self.tree_draw))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="在小批中反查BOM", command=lambda: self.find_parent_GUI(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="在小批中查询子零件", command=lambda: self.find_child_GUI(self.tree_code))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="在设计BOM中反查BOM", command=lambda: self.find_parent_GUI(self.tree_code,'DESIGN'))
        self.menu1.add_separator()
        self.menu1.add_command(
            label="在设计BOM中查询子零件", command=lambda: self.find_child_GUI(self.tree_code,'DESIGN'))
        self.menu1.add_separator()
        self.menu1.add_command(label="在当前BOM中查询", command=self.find_treebom_GUI)
        self.menu1.add_separator()
        self.menu1.add_command(label="打开图纸", command=self.open_draw_GUI)        
        self.menu1.add_separator()        
        self.menu1.add_command(label="导出列表",command=self.tree_save)

    def menu_bar(self,):   # 定义菜单栏   
        m_bar = tk.Menu(self.root)  # 创建菜单组

        m_file = tk.Menu(m_bar, tearoff=0)  # 创建2级菜单组
        m_file.add_separator()
        m_file.add_command(label='导入ERP BOM',command=lambda: self.read_bom_GUI(tp='BATCH'))
        m_file.add_separator()        
        m_file.add_command(label='导入设计BOM',command=lambda: self.read_bom_GUI(tp='DESIGN'))
        m_file.add_separator()
        m_file.add_command(label='更新物料库',command=self.read_code_GUI)
        m_file.add_separator()
        m_file.add_command(
            label='临时读取设计BOM', command=lambda: self.read_bom_GUI(tp='DESIGN',type1='TEMP'))
        m_file.add_separator()
        # mabr上添加一个标签,链接到file_m
        m_bar.add_cascade(label='读取EXCEL文件', menu=m_file)

        m_cost = tk.Menu(m_bar, tearoff=0)
        m_cost.add_separator()
        m_cost.add_command(label='导入成本文件', command=self.read_cost_GUI)
        m_cost.add_separator()
        m_cost.add_command(label='变动成本', command=self.view_changed_cost)
        m_cost.add_separator()
        m_cost.add_command(label='重算组合件成本', command=self.recalc_cost)
        m_cost.add_separator()        
        m_cost.add_command(label='查看当前物料成本', command=self.tree_add_cost)

        m_bar.add_cascade(label='成本', menu=m_cost)

        m_view = tk.Menu(m_bar, tearoff=0)  # 创建2级菜单组
        root_b = tk.StringVar()
        
        for root,name in batch_root.items():
            # 单选菜单整组有一个variable，每个选项都有一个value。当被选中时，该菜单的value就会赋值给variable。
            m_view.add_radiobutton(label='小批BOM:'+root +' '+ name, value=root, variable=root_b,
                                   indicatoron=False, command=lambda: self.find_child_GUI(root_b.get()))
            m_view.add_separator()
        for root,name in design_root.items():
            # 单选菜单整组有一个variable，每个选项都有一个value。当被选中时，该菜单的value就会赋值给variable。
            m_view.add_radiobutton(label='设计BOM:'+root +' '+ name, value=root, variable=root_b,
                                   indicatoron=False, command=lambda: self.find_child_GUI(root_b.get(),db='DESIGN'))
            m_view.add_separator()    

        m_bar.add_cascade(label=' 查看导入的BOM ', menu=m_view)        

        m_path = tk.Menu(m_bar, tearoff=0)
        draw_p = tk.StringVar()
        for key, item in path_root.items():
            m_path.add_radiobutton(label=key, value=item, variable=draw_p,
                                   indicatoron=False, command=lambda: startfile(draw_p.get()))
            m_path.add_separator()    
        m_path.add_separator()
        m_path.add_command(label='编辑产品路径', command=self.edit_path_GUI)
        
        m_bar.add_cascade(label='产品路径', menu=m_path)

        m_tool = tk.Menu(m_bar, tearoff=0)
        m_tool.add_separator()
        m_tool.add_command(label='检查Excel编码', command=lambda:self.check_excel_GUI(tp='CHECK'))
        m_tool.add_separator()
        m_tool.add_command(label='添加部件数量', command=lambda:self.check_excel_GUI(tp='QTY'))
        m_bar.add_cascade(label='EXCEL工具', menu=m_tool)

        self.root.config(menu=m_bar)  # 把mbar菜单组 配置到窗体;
#-----------------以下窗口动作触发------------------------------      
    def R_click_en1(self, event):   # 输入框绑定动作
        self.menu_eny1.post(event.x_root, event.y_root)   # 在事件坐标处,弹出对应的菜单

    def R_click_tree(self, event):   # 鼠标右键绑定的动作，该程序通过前面的bind 和右键绑定在一起
        iid = self.tev.identify_row(event.y)   # 返回事件发生时鼠标坐标对应的行
        if iid:   # 如果鼠标所在是空,则不执行右键动作
            self.tev.selection_set(iid)    # 当右键时选中目前鼠标所在的行
            self.tree_code = self.tev.item(self.tev.selection(), 'values')[0]
            self.tree_draw = self.tev.item(self.tev.selection(), 'values')[1]
            name = self.tev.item(self.tev.selection(), 'values')[2]
            self.x_root = event.x_root
            self.y_root = event.y_root
            if self.tree_code == '-':
                self.tree_code = self.tree_draw + name
            else:
                self.tree_code = self.tree_code.replace('old', '')
            self.menu1.post(event.x_root, event.y_root)

    def en1_enter(self, event=None):  #和事件绑定的函数,在事件触发时,会自动给一个event参数,所有定义时必须加上
        self.find_code_GUI('BATCH')

#-----------------以下窗口动作函数--------------------------------        
    def find_code_GUI(self,db='BATCH'):
        x = self.en1.get()
        if x in ('', ' ', None):
            pass
        elif len(x) < 2:
            self.lab_r.set('符合条件物料太多，请补充信息')
        x = x.replace('\n', '')  #去除换行符  
        txt={'BATCH':'小批','DESIGN':'设计'}
        rst=find_code(x,db=db)
        if 'code' in rst:
            rst['code'].sort(key=lambda x: x[0])            
            rst['code'] = [(1,) + x for x in rst['code']]
            
            self.lab_r.set(str(x)+' 在{0}库的查询结果:'.format(txt[db]))
            self.tree_out(rst['code'],type='CODE')
        else:
            self.lab_r.set(str(x) + ' 在{0}库的未找到相关信息:'.format(txt[db]))

    def find_parent_GUI(self, x,db='BATCH'):        
        rst = find_parent_bom(x,db)
        if 'bom' in rst:
            self.lab_r.set(str(x)+' 的反查结果:')
            self.tree_out(rst['bom'], tar=x)
        else:
            self.lab_r.set('没有BOM中使用此物料：%s'%str(x))

    def find_child_GUI(self, x,db='BATCH'):        
        rst=find_child_bom(x,db)
        if 'bom' in rst:            
            self.lab_r.set(str(x)+' 的子项结构查询结果')
            self.tree_out(rst['bom'])
        else:
            self.lab_r.set('%s 物料没有子零件'%str(x))

    def read_bom_GUI(self,tp='BATCH',type1=''):
        file_name = tk.filedialog.askopenfilename(title='打开BOM文件',filetypes=[('xlsx', '*.xlsx'),])        
        rst = read_design_BOM(file_name,type=tp)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'itemerror' in rst:
            self.lab_r.set('表格内有如下错误：')
            self.tree_out(rst['itemerror'])
        elif 'bom' in rst:
            if type1!='TEMP':
                wr2 = tk.messagebox.askquestion(message='对已存在的BOM层次,进行覆盖还是跳过?')
                if wr2=='yes':
                    type1 = 'W'
            root,rst_t1 = update_to_bom_db(rst['bom'], db=tp, type=type1)
            self.find_child_GUI(root,db=tp)
            self.lab_r.set(rst_t1)

    def read_cost_GUI(self,):
        file_name = tk.filedialog.askopenfilename(title='打开成本文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = read_design_BOM(file_name,type='COST')
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'itemerror' in rst:
            self.lab_r.set('表格内有如下错误：')
            self.tree_out(rst['itemerror'])
        elif 'bom' in rst:            
            rst1 = update_to_cost_db(rst['bom'])
            if 'error' in rst1:
                self.lab_r.set(rst1['error'])
            elif 'change' in rst1:
                self.lab_r.set(rst1['info'])
                self.tree_out(rst1['change'])            

    def read_code_GUI(self,):
        self.code_mod = []
        popcode = POP_readcode(self, self.root)
        self.root.wait_window(popcode.pop)

        if self.code_mod:
            self.tree_out(self.code_mod,type='CODE')

    def check_excel_GUI(self,tp='CHECK'):   #查找excel表编码
        file_name = tk.filedialog.askopenfilename(title='打开BOM文件',
                                                  filetypes=[('xlsx', '*.xlsx'),])
        
        rst = read_design_BOM(file_name,type=tp)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'itemerror' in rst:
            self.lab_r.set('表格内有如下错误：')
            self.tree_out(rst['itemerror'])
        else:
            self.lab_r.set('更新完成，改动如下：')
            self.tree_out(rst['bom'])

    def find_treebom_GUI(self,):
        f = askstring("在列表中查找", "请输入要查询的内容")
        f=f.upper().strip()  # 转大写，去收尾空格
        f = f.replace('\n', '')  #去掉换行符
        
        if f:
            self.tree_out(self.tree_bom,tar=f)          

    def tree_out(self, bom,type='BOM',tar=''):  # 向treeview中写入列表内容
        # BOM格式[0层次,1编码,2图号,3名称,4数量,5本层数量]
        # 设计BOM格式[0层次,1编码,2图号,3名称,4数量,5本层数量,6材料,7备注]
        # 物料格式[0层次,1编码,2图号,3名称,4日期]
        # 成本格式[0层次,1编码,2图号,3名称,4材料成本,5人工成本,6管理成本,7总成本]
        #self.lab_r.set(self.target)
        def set_tree_title(type):
            tree_title = {
            'CODE': ('序号', '编码', '图号', '名称', '材料', '重量', '备注',),
            'CODE-COST': ('序号', '编码', '图号', '名称', '材料成本', '人工成本', '管理费用', '单件成本', '更新日期',),
            'BOM': ('序号', '编码', '图号', '名称', '数量', '部件数量', '材料', '重量', '备注',),
            'BOM-COST': ('序号', '编码', '图号', '名称', '数量', '部件数量', '单件成本', '合计成本', '更新日期',),
            }
            title_width = {
                '序号':100,'编码':100, '图号':100, '名称':250, '材料':80, '重量':60,'数量':40, '部件数量':40, '备注':80,'材料成本':50, '人工成本':50, '管理费用':50, '单件成本':60,'合计成本':60, '更新日期':150,
            }
            cols = (str(x) for x in range(1, len(tree_title[type]) + 1))           

            if type in tree_title:
                self.tree_col=tree_title[type]
                for n, name in enumerate(tree_title[type]):
                    if n == 0:
                        n='#'+str(n)
                    self.tev.heading(str(n), text=name)
                    self.tev.column(str(n), width=title_width[name])
                    
        def set_tags(item1):
            s = self.tev.item(item1, 'values')
            if tar in str(s) or (s[0]=='-' and tar==s[1]+s[2]):                
                if len(s)>4:
                    rst.append(s[4])
                else:
                    rst.append(s[0])

                self.tev.item(item1, tag='tar')
                p1 = self.tev.parent(item1)
                p2 = self.tev.parent(p1)
                p3 = self.tev.parent(p2)
                p4 = self.tev.parent(p3)
                self.tev.item(p1, open=True)
                self.tev.item(p2, open=True)
                self.tev.item(p3, open=True)
                self.tev.item(p4, open=True)

        self.tree_bom = bom
        self.tree_type = type
        self.tree_col=[]
        for item in self.tev.get_children():  # 对treeview进行清空
            self.tev.delete(item)

        set_tree_title(type)
        rst=[]
        lv = {0: ''}
        a = bom[0][0] - 1
        order_n = [0, 0, 0, 0, 0, 0, 0, 0, 0]

        self.tev.tag_configure('tar', background='blue')
        for key in bom:
            order_n[key[0] + 1:8] = 0,0,0,0,0,0,0,0
            order_n[key[0]] += 1

            i = key[0] - a
            lv[i] = self.tev.insert(lv[i - 1],'end',text=str(order_n[key[0]]),values=key[1:])
            if tar:
                set_tags(lv[i])
            elif i == 1:
                self.tev.item(lv[i], open=True)

        #self.tev.tag_configure('tar', foreground='blue',background='#eeeeff', font='微软雅黑')

        if rst:
            t1='%s 在列表中共出现: %d 次;'%(tar,len(rst))            
            try:
                n=sum([float(x) for x in rst])
                t1 += ' 总用量为: %d ;' % (n)
            except:
                pass
            self.lab_r.set(t1) 
        elif tar:
            t1='%s 在列表中未找到'% tar
            self.lab_r.set(t1)                              

    def tree_save(self,):
        file = tk.filedialog.asksaveasfilename(defaultextension=".xlsx",title='保存文件',
                                                  filetypes=[('xlsx', '*.xlsx')])
        rst = save_to_excel(file, self.tree_bom, self.lab_r.get(),self.tree_col)
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        else:
            self.lab_r.set('已成功导出到文件：' + file)

    def tree_add_cost(self,):  # 在当前显示的物料后添加成本数据
        def parent_cost():  #对装配体成本为0的进行重新计算
            cost_lv = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}
            complete_lv={1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1}
            for n, item in enumerate(cost_bom[::-1]):
                x=0
                if (item[1]=='-' or item[1][:3] in ('24R','28R')) and cost_lv[item[0] + 1] != 0:
                    c = round(cost_lv[item[0] + 1],2)
                    ct = round(c * item[5], 2)
                    m = len(cost_bom) - n
                    if item[6] == '':  # 当原成本为0时
                        if complete_lv[item[0] + 1]:
                            ss = ''
                        else:
                            ss=' (*)'
                        cost_bom[m - 1] = item[:6] + (c, ct) + ('子件计算'+ss,)
                        x=c*item[4]
                    elif abs(item[6] - cost_lv[item[0] + 1]) > 1:  #原成本存在时和子件计算成本进行比较
                        if complete_lv[item[0] + 1]:
                            ss = ''
                        else:
                            ss='(*) '
                        cost_bom[m - 1] = item[:6] + (c, ct) + (ss+item[8] + '：' + str(item[6]),)
                        x = c * item[4]                    
                            
                if x:
                    cost_lv[item[0]] += x
                elif isinstance(item[6], (int, float)):
                    cost_lv[item[0]] += item[6] * item[4]
                else:   #当有零件无成本时,标记该层次的成本不完整
                    complete_lv[item[0]]=0

                for i in range(item[0] + 1, 8):
                    cost_lv[i] = 0
                    complete_lv[i]=1

        def total_cost():
            tot_m=tot_l=tot_e=0
            for item in cost_bom:
                if item[0] == 2:
                    tot_m+=item[7]
                    
        cost_bom = []        
        if self.tree_type=='BOM':
            for item in self.tree_bom:
                if item[1] in all_cost:
                    tot = all_cost[item[1]][3]
                    d=all_cost[item[1]][4]
                    item = tuple(item[:6]) + (round(tot,2), round(tot * item[5],2),d)
                else:
                    item = tuple(item[:6]) + ('','')
                cost_bom.append(item)
            parent_cost()

            self.tree_out(cost_bom,type='BOM-COST')
        elif self.tree_type == 'CODE':
            for item in self.tree_bom:
                if item[1] in all_cost:
                    cm = all_cost[item[1]][0]
                    cl = all_cost[item[1]][1]
                    ce = all_cost[item[1]][2]
                    ct = all_cost[item[1]][3]
                    d=all_cost[item[1]][4]
                else:
                    cm=ct=cl=ce=d=''
                item = item[:4] + (ct,cm,cl,ce,d)
                cost_bom.append(item)
            self.tree_out(cost_bom, type='CODE-COST')
        self.lab_r.set('物料成本查询如下 ( * 表示子件成本不完整):')            

    def view_changed_cost(self,): # 查看变动的成本
        rst = load_old_cost()
        if rst:
            self.lab_r.set('成本变动过的物料如下:')
            self.tree_out(rst, 'CODE-COST')
        else:
            self.lab_r.set('库中物料成本没有发生变化')

    def recalc_cost(self,):
        rst = reclac_parent_cost()
        if 'error' in rst:
            self.lab_r.set(rst['error'])
        elif 'change' in rst:
            self.tree_out(rst['change'], 'CODE-COST')

    def edit_path_GUI(self,):
        edit_root_path(self, self.root)
        #self.root.wait_window(p.pop)

    def open_draw_GUI(self,):
        def open_file(filepath):
            try:
                startfile(filepath)
            except Exception as ex:
                self.lab_r.set(str(ex))

        def meun_drawpath():
            self.menu_path_sect = tk.Menu(self.root, tearoff=0)
            path_1 = tk.StringVar()
            for item in rst:
                self.menu_path_sect.add_radiobutton(label=item[1], value=item[1], variable=path_1,
                                   indicatoron=False, command=lambda: open_file(path_1.get()))
                self.menu_path_sect.add_separator()

        rst = {}        
        if self.tree_draw != '-' and 'GB' not in self.tree_draw:
            rst = find_db('draw', self.tree_draw, 'drawPATH')
            if rst:
                if len(rst) == 1:
                    open_file(rst[0][1])
                else:
                    meun_drawpath()
                    self.menu_path_sect.post(self.x_root, self.y_root)
            else:
                self.lab_r.set('没有找到对应图纸')
       
#-----------------以下主程序函数------------------------------ 
def find_parent_bom(f,db='BATCH'):    # 根据编码反查使用的BOM
    # 在字典每个值的里面查找编码,找到后将对应的key,再作为编码进行同样查找,直到key=index
    def find_parent(x, n):
        parent[n]=''
        for key, item in all_bom.items():
            for code in item:
                if x in code:
                    parent[n] = [n] + list(code)
                    find_parent(key, n - 1)
                    break
        if not parent[n] and x!=f:
            parent[n] = [n] + [x, 1]
            if x not in father_bom_dict:
                father_bom_dict[x] = []            
            num = n - 1
            p1=[]
            for i in range(n, 10):                               
                p1.append([parent[i][0] - num, parent[i][1], parent[i][2]])
            father_bom_dict[x].append(p1[:])

    def fmt_father_bom():    # 去除相同的父项,并由字典转化为列表形式
        for key in father_bom_dict:
            lvcode = {}
            father_bom_dict[key].sort()
            for items in father_bom_dict[key]:
                for item in items:
                    if lvcode.get(item[0], 'NA') != item[1]:
                        lvcode[item[0]] = item[1]
                        for n in range(item[0] + 1, 7):
                            lvcode[n] = ''
                        bom.append([item[0], item[1], item[2]])

    def father_bom_total(x):
        # 计算反查物料在顶层的总用量,和本层用量
        lv_num = {0: 1, 1: 1}
        code_num = [0]
        root_index = 0

        for n, item in enumerate(bom):
            lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
            item.append(lv_num[item[0]])

            if item[1] == x:
                code_num.append(lv_num[item[0]])

            if item[0] == 1 or n == len(bom) - 1:
                bom[root_index][3] = (sum(code_num))
                code_num = [0]
                lv_num = {0: 1, 1: 1}
                root_index = n

    bom = []
    parent={}
    father_bom_dict = {}
    rst={}
    if db.upper() == 'BATCH':
        all_bom = all_batch_bom
    elif db.upper() == 'DESIGN':
        all_bom = all_design_bom

    find_parent(f, 9)    
    if father_bom_dict:
        fmt_father_bom()
        father_bom_total(f)
        if db.upper() == 'BATCH':
            for n,item in enumerate(bom):
                code,draw, name = get_code_info(item[1])
                bom[n]=(item[0],code,draw,name,item[2],item[3])
                
        elif db.upper() == 'DESIGN':
            for n,item in enumerate(bom):
                code,draw,name,metal,weight,remark=get_designcode_info(item[1])
                if code == draw + name:
                    code = '-'
                bom[n] = (item[0], code, draw, name, item[2], item[3], metal, weight, remark)

        rst['bom']=bom
    return rst

def find_code(f,item='ALL',exact=False,db='BATCH'):  # 根据输入内容查找物料，item为查找的字段，exact为TRUE则准确匹配，false可以使用通配符*；
    rst_code = []
    
    f=str(f)
    f=f.upper().strip()  # 转大写，去首尾空格
    f = f.replace('\n', '')  #去掉换行符
    rst = {}

    if db.upper() == 'BATCH':
        all_code = all_batch_code
    elif db.upper() == 'DESIGN':
        all_code = all_design_code
        #code_db.update()
        
    if f in all_code:
        rst_code.append(all_code[f][:])

    elif exact:     # 准确匹配，要求完全相同
        for item in list(all_code.values()):                
            if f in str(item):
                rst_code.append(item[:])
    else:
        f = f.replace('.', '\.')     # 点保持原样，不要被改成通配符        
        f = f.replace('*', '.*')    # 将windows习惯用法的 * 转换为python中的 .*
        f = f.replace(' ','.*')       # 字符中间空格，改为通配符
        x = re.compile(f)            # 使用正则表达式中通配符进行查询
        for item in list(all_code.values()):            
            if x.search(str(item)):
                rst_code.append(item[:])   #要对元素进行添加,而不是整个地址引用,那样会造成原列表被修改
    
    if rst_code:        
        if db.upper() == 'DESIGN':  #对于设计BOM中编码=图号+名称的,编码用'-'代替
            for n,item in enumerate(rst_code):
                if item[0] == item[1] + item[2]:                    
                    rst_code[n]=('-',)+item[1:]                  
        rst['code']=rst_code    #[编码，图号，名称].
    return rst

def find_child_bom(f,db='BATCH'):      # 根据编码查找子零件
    def bom_total():
        # 计算反查物料在顶层的总用量,和本层用量
        lv_num = {0: 1, 1: 1}       
        for item in bom:
            lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
            item.append(lv_num[item[0]])
    
    def find_child(x, n):
        if x in all_bom:
            for item in all_bom[x]:
                bom.append([n]+list(item))
                find_child(item[0],n+1)
    
    bom = []
    rst = {}
    if db.upper() == 'BATCH':
        all_bom = all_batch_bom
    elif db.upper() == 'DESIGN':
        all_bom = all_design_bom

    if f in all_bom:
        bom.append([1,f,1])
        find_child(f, 2)
        bom_total()
        if db.upper() == 'BATCH':
            for n,item in enumerate(bom):
                code,draw,name = get_code_info(item[1])
                bom[n]=(item[0],code,draw,name,item[2],item[3])
                
        elif db.upper() == 'DESIGN':
            for n,item in enumerate(bom):
                code,draw,name,metal,weight,remark=get_designcode_info(item[1])
                if code == draw + name:
                    code = '-'
                bom[n]=(item[0],code,draw,name,item[2],item[3],metal,weight,remark)

        rst['bom'] = bom     
    return rst

def get_code_info(s):  # 根据编码返回图号和名称    
    sa = s.replace('P', '')
    if sa in all_batch_code:
        return all_batch_code[sa]
    elif sa in batch_root:
        return sa,'-',batch_root[sa]
    else:
        return '-', '-', '-'

def get_designcode_info(s):  # 根据编码返回图号和名称     
    if s in all_design_code:
        return all_design_code[s]          
    elif s in all_batch_code:
        return all_batch_code[s] + ('-', 0, '-')
    elif s in design_root:
        return s, '-', design_root[s], '-', 0, '-'
    elif s in batch_root:
        return s, '-', batch_root[s], '-', 0, '-' 

def read_design_BOM(file,type='BATCH'):  # 统一读取各种bom文件,输出格式为统一的
    # 先判断属性列和root,再读取并按统一格式生成列表:(除了名字和编码以外的列都允许不存在,由后续程序进行判断)
    # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]
    # 第一行是表头，有属性的为字段，无属性的为‘-’
    def get_col(wsheet):
        col.clear()
        lable = {
            '级别': 'lv',
            '层次': 'lv',
            '序号': 'lv',
            '编码': 'code',
            '子件编码': 'code',
            '新编码':'code',
            '图号': 'draw',
            '代号': 'draw',                      
            '名称': 'name',
            '子件名称': 'name',
            '描述':'name',
            '用量': 'num',
            '数量': 'num',
            '基本用量': 'num',
            '使用数量' :'num',           
            '材料': 'metal',
            '单重':'weight',
            '备注': 'remark',
            '材料成本': 'cost_mt',
            '人工成本': 'cost_lb',     
            '费用成本': 'cost_exp', 
        }
        lable_t={}
        for key,item in lable.items():
            if item in title:
                lable_t[key] = item
                
        str_row = 1
        for row in wsheet.values:
            str_row += 1
            if str_row > 10:
                rst['error']='前10行找不到基本的属性列:'+str(title)
                break         
            for c, value in enumerate(row):
                for key in lable_t:
                    if isinstance(value, str) and key == value.replace(' ', ''):
                        if lable_t[key] in col:
                            if lable_t[key] == 'lv':  # 检查属性是否重复
                                col['lv_end'] = c
                            else:                        
                                rst['error']='%s 的属性列重复'% key
                                col.clear()
                                return
                        else:
                            col[lable_t[key]] = c

            if title[0] in col and 'name' in col:
                col['str_row'] = str_row
                if 'lv' in col:
                    if 'num' not in col:
                        rst['error'] = '找不到数量列'
                    
                    if 'lv_end' not in col:
                        col['lv_end'] = col['lv']
                break
            else:
                col.clear()
    
    def get_lv(lvs):
        if not excel_bom:   #先要找到ROOT才能继续
            if isinstance(lvs[0], str) and lvs[0].upper() == 'ROOT':                
                return 1
            else:                
                return ' 行没有ROOT'
        elif len(lvs)==1:    #针对层次只有1列：一种是'+++'，另一种是本身只有一层1，2，3，4
            if isinstance(lvs[0], str):                
                lv = len(lvs[0])
            elif isinstance(lvs[0], int):
                lv = 2
            else:
                lv = ' 缺少层次'
            return lv
        else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
            lv_1 = []
            for col, n in enumerate(lvs):
                if isinstance(n, int):
                    lv = col+1
                    lv_1.append(lv)
            if len(lv_1) == 1:
                return lv
            elif len(lv_1) > 1:
                return ' 行层次重复'
            else:
                return ' 行缺少层次'

    def check_excel_code(code, draw,row_num):  #对编码进行检查
        code1 = check_code(code, draw)        
        if type=='CHECK':
            if code1 != code:
                fill_blue = PatternFill('solid',fgColor='EFBF00')                
                wsheet.cell(row=row_num, column=col['code'] + 1).value = code1
                wsheet.cell(row=row_num, column=col['code'] + 1).fill = fill_blue
        elif type == 'DESIGN':
            pass
        return code1.replace('old', '')

    def write_QTY(lv, num, row_num):        
        for n in range(lv + 1, 7):
            lv_num[n]=0
        lv_num[lv] = num * lv_num[lv - 1]
        wsheet.cell(row=row_num, column=col['num'] + 2).value = lv_num[lv]

    def read_item(wsheet):
        def fmt_str(x):
            if x:
                x=str(x)
                x = x.replace(' ', '')
                if x:
                    return x.upper()
                else:
                    return '-'
            else:
                return '-'
        def fmt_num(x):
            if not x:
                return 0
            elif isinstance(x, int):
                return x            
            else:
                try:                
                    return round(float(x), 2)
                except:
                    return x

        row_num = col['str_row'] - 1
        chd_lv = 0
        for row in wsheet.iter_rows(min_row=col['str_row'], values_only=True):
            item={}
            row_num += 1
            for key in title:
                if key in col:
                    if key == 'lv':                        
                        item[key] = get_lv(row[col['lv']:col['lv_end'] + 1])
                    elif key in ('num','weight','cost_mt', 'cost_lb', 'cost_exp'):
                        item[key] = fmt_num(row[col[key]])
                    else:
                        item[key] = fmt_str(row[col[key]])
                else:
                    item[key] = '-'

            if item['name'] == '-':  #跳过空行
                continue

            if 'lv' in item:
                if isinstance(item['lv'], int):
                    if len(excel_bom)==1:   # 第二行的层次必须是2, 所以要根据第二行的层次来确定一个整体层次的调整系数
                        chd_lv = 2 - item['lv']                            
                    item['lv'] = item['lv'] + chd_lv
                    last_lv = item['lv']
                    if item['lv'] > 1 and item['lv'] > last_lv + 1:  # 检查层次是否连续
                        item_error.append('第 ' + str(row_num) + ' 行层次和上层脱节')
                else:
                    item_error.append('第 ' + str(row_num) + item['lv'])
                                                        
            if 'num' in item:
                if isinstance(item['num'], (int, float)):
                    if type == 'QTY' and not item_error:
                        write_QTY(item['lv'],item['num'], row_num)                
                else:
                    item_error.append('第 ' + str(row_num) + '行没有数量或格式不对')

            if 'code' in item:   #对编码格式进行检查,并
                if type in ('DESIGN','CHECK') and not item_error:
                    item['code']=check_excel_code(item['code'],item['draw'],row_num)

                if re.match(rule['code'], item['code']) or re.match(rule['root'], item['code']) or  re.match(rule['rootnew'], item['code']):
                    pass
                elif item['code'] == '-' and type in ('DESIGN', 'CHECK'):  #有时候允许为-
                    pass
                elif type == 'CODE':  #当读取物料库时，允许编码错误，其它情况下提示错误
                    continue
                elif type == 'TEMP':  #当TEMP时,不检查编码格式
                    pass
                else:
                    item_error.append('第 ' + str(row_num) + '行编码格式不对')

            if not item_error:
                t=[]
                for k in title:
                    t.append(item[k])
                excel_bom.append(t)
            
    excel_bom = []
    col = {}
    rst={}
    skip_sheet = ''
    item_error=[]
    m=True
    if type in ('BATCH','DESIGN','TEMP'):
        title = ('lv', 'code', 'draw', 'name', 'num', 'metal', 'weight', 'remark')
    elif type == 'CODE':
        title = ('code', 'draw', 'name')
    elif type == 'COST':
        title = ('code', 'name', 'cost_mt', 'cost_lb', 'cost_exp')
    elif type == 'CHECK':
        m = False
        title = ('code', 'draw', 'name')
    elif type == 'QTY':
        lv_num = {0: 1, 1: 1}
        m = False
        title = ('lv', 'name','num')

    try:
        wbook = xl.load_workbook(file, read_only=m)
    except Exception as ex:        
        rst['error'] = '文件读取失败：'+str(ex)
        return rst
    
    if type == 'CODE':
        names=wbook.sheetnames
    else:
        names = [wbook.active.title]
           
    for sname in names:        
        wsheet = wbook[sname]
        get_col(wsheet)
        if col and 'error' not in rst:
            if type == 'CHECK':
                wbook.copy_worksheet(wsheet)    #创建一个原工作表的备份
            elif type == 'QTY':
                wbook.copy_worksheet(wsheet)    #创建一个原工作表的备份
                wsheet.insert_cols(col['num']+2)   #在数量列后面插入一列,inset会插入在前面,cols时列从1开始，而rows数组从0开始

            read_item(wsheet)
        else:
            skip_sheet += wsheet.title
        
    if 'error' in rst:
        rst['error'] = wsheet.title + ' 表的' + rst['error']    
    elif item_error:
        item_error=[[1,x] for x in item_error]
        rst['itemerror'] = item_error        
    elif type in ('CHECK','QTY'):  #如果是检查编码模式，则将写入的结果保存到excel表
        if excel_bom:            
            try:
                wbook.save(file)
            except Exception as ex:
                rst['error'] = 'EXCEL文件写入失败：' + str(ex)
        else:
            rst['error'] = '所有编码未改动'
    else:
        head=[]
        for key in title:
            if key in col:
                head.append(key)
            else:
                head.append('-')         
        excel_bom.insert(0,head)
        rst['bom'] = excel_bom
                
    if skip_sheet:
        rst['skip'] = skip_sheet

    wbook.close()
    return rst  #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]

def check_code(code,draw):   # 检查编码
    # 如果有编码,如果没有编码且有图号,则根据图号去编码库查找
    
    if not re.match(rule['code'],code):
        if draw != '-' and 'GB' not in str(draw):
            rst=find_code(draw, item='ALL', exact=True)
            if 'code' in rst and len(rst['code'])==1:
                code= rst['code'][0][0]                
    elif code in all_batch_code:
        code= all_batch_code[code][0]
    return code

def update_to_bom_db(excel_bom,db='BATCH',type='W'): #将已读取的设计BOM写入设计BOM库,分别创建designcode和designbom库
    #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注]
    # 把设计BOM按小批样式，分物料库和BOM库，用一个自增数字作为代码进行唯一区别
    #designcode结构：code:[code,draw,name,metal,weight,remark]
    #designbom结构：[code:[[code1,num1],        ]           ]
    def creat_bom():  #对没有编码的,以图号和名称为编码         
        for item in excel_bom:
            if item[1]=='-':     #对于找不到编码的，用图号+名称作为编码
                item[1] = item[2] + item[3]
            ebom.append(tuple(item))

    def creat_new_code():  # 筛选出需要保存到designCODE的物料        
        dcode = (('15R', '16R', '24R', '28R')+
                ('C01', 'C02', 'C03', 'C04', 'C05', 'C06','C07', 'C08')+
                ('E01', 'E02', 'E03', 'E04', 'E05', 'E06', 'E07', 'E08'))

        for item in ebom:
            code = item[1]
            draw = item[2]
            name = item[3]            
            metal = item[5]
            weight= item[6]
            remark = item[7]
            item = (code, draw, name, metal, weight, remark)

            if code not in all_design_code:
                if code in all_batch_code:
                    if code[:3] in dcode:
                        new_code[code] = item
                else:
                    new_code[code] = item
            elif type=='W':
                if all_design_code[code] != item:
                    new_code[code] = item

    def creat_bom_dict():  #创建读取物料的设计BOM的字典
        bom={}         
        lv_code = {}
        ex_code_lv = 10
        
        for n in ebom:
            code = n[1]
            lv = n[0]
            num = n[4]
            if lv > ex_code_lv:
                continue
            elif lv>1:
                bom[lv_code[lv - 1]].append((code, num))

            if code in bom:
                ex_code_lv = lv
            else:
                lv_code[lv] = code
                bom[code] = []
                ex_code_lv = 10        

        for key, item in bom.items():
            if item:
                ebom_dict[key] = item

    def creat_new_bom(): #将读取的设计BOM和已有的设计BOM结构进行对比，新的添加到new_bom，有变动的物料添加到old，然后到designBOM删除结构
        for key, item in ebom_dict.items():
            if key not in all_bom:
                for code in item:
                    new_bom.append((key, code[0], code[1]))
            elif type=='W': #因为内部元素顺序可能不同，所以要转为set，再进行对比                
                if set(item) != set(all_bom[key]):
                    old.append(key)
                    for code in item:
                        new_bom.append((key, code[0], code[1]))

    if not ('lv' and 'code' and 'name' and 'num') in excel_bom[0]:
        return 'BOM缺少必要的属性列（层次，编码，名称，数量）'
    else:
        del excel_bom[0]  #删除BOM表里面的属性头

    

    ebom = []
    ebom_dict={}
    new_code = {}   #新的物料,只有当是设计BOM时才创建;用于写入设计物料库
    new_bom = []    #新的BOM层次,用于写入数据库
    old = []        #有变动的物料,用于在数据库中删除
    txt={'BATCH':'小批','DESIGN':'设计'}
    db=db.upper()
    if db == 'BATCH':
        all_bom = all_batch_bom
        dbsheet_bom = 'batchBOM'
        dbsheet_code = 'code'
        
    elif db == 'DESIGN':
        all_bom = all_design_bom
        dbsheet_bom = 'designBOM'
        dbsheet_code = 'designCODE'
        
        creat_bom()
        creat_new_code()

    creat_bom_dict()
    creat_new_bom()
    root = excel_bom[0][1]
    draw = excel_bom[0][2]
    rootname = excel_bom[0][3]
    
    if type=='TEMP':  # 如果是临时性读取bom,则不进行数据库文件的保存
        all_design_code.update(new_code)
        all_bom.update(ebom_dict)

        return root,'临时读取的BOM为: '
    try:        
        if new_code:
            insert_db(new_code.values(), dbsheet_code)
            all_design_code.update(new_code)
        if re.match(rule['root'],root) or re.match(rule['rootnew'],root):
            if db == 'BATCH':
                insert_db(((root, draw, rootname),), 'code')
            insert_db(((root, rootname, db),), 'root')  #将root信息写入root 
            load_root_db()
        if old:
            old=old+['code']
            remove_db(tuple(old), sheet=dbsheet_bom)        
        if new_bom:            
            insert_db(new_bom, dbsheet_bom)

        all_bom.update(ebom_dict)
        conn.commit()        
        return root,root + '已成功写入{0} 库'.format(txt[db])
    except Exception as ex:
        conn.rollback()
        return root,'{0}数据库写入失败,已撤销:'.format(txt[db]) + str(ex)

def update_to_code_db(excel_bom):   #根据读取的物料信息,将变动过的写入数据库    
    #[0编码,1图号,2名称]
    def old_item_check():  # 查找旧编码        
        keys=[]
        for key in all_batch_code.keys():  # 取出字典中所有的key,组成list，然后排序进行比较
            if key.startswith('15R') or key.startswith('16R') or key.startswith('24R') or key.startswith('28R'):
                keys.append(key)
        
        keys.sort()
        # 物料按编码排序,对任一编码和下一个编码进行比对,如果除了最后一位相同,且比较小,则认为是旧编码
        for x, a in enumerate(keys):            
            if 'old' in all_batch_code[a][0]:
                continue            
            a = a.replace('P', '')
            b = keys[x + 1]
            if b[-1] == 'P':
                old.append(a)
                all_batch_code[a]= (all_batch_code[a][0]+'old',)+all_batch_code[a][1:3]
                    
            elif a[:-1] == b[:-1] and a[-1]<b[-1]:
                old.append(a)
                all_batch_code[a]= (all_batch_code[a][0]+'old',)+all_batch_code[a][1:3]
                    
            if x == len(keys) - 2:
                break

    def creat_code_dict():
        for item in excel_bom:
            if isinstance(item[0], str) and item[0]!='code':
                excel_code[item[0]] = tuple(item)
            
    def creat_new():
        for key,item in excel_code.items():            
            if key not in all_batch_code:
                change.append(item)
            elif item != all_batch_code[key] and 'old' not in all_batch_code[key][0]:                
                change.append(item)
                mod_list.append([1] + list(item) + ['新修改'])
                mod_list.append([2]+list(all_batch_code[key])+['原来的'])

    rst = {}
    excel_code={}
    change = []    
    old=[]
    mod_list=[]
    
    if not ('code' and 'name') in excel_bom[0]:
        return '编码表缺少必要的属性列（编码,名称）'
    else:
        del excel_bom[0]  #删除BOM表里面的属性头
    
    creat_code_dict()
    creat_new()
    all_batch_code.update(excel_code)
    old_item_check()
    
    if mod_list:
        rst['mod'] = mod_list
    try:
        if old:
            set_old_item(tuple(old), 'code')
        if change:
            rst['new'] = len(change)-len(mod_list)/2
            insert_db(change, 'code')
        conn.commit()
    except Exception as ex:
        conn.rollback()
        rst['error']='写入文件失败'+str(ex)

    return rst

def update_to_cost_db(excel_bom):  #根据读取的BOM,找出成本变动过的,写入数据库
    #[0编码,1名称,2材料成本,3人工成本,4管理成本] 
    def creat_cost():
        for item in excel_bom:
            metal = item[2] if isinstance(item[2],(int,float)) else 0
            labor = item[3] if isinstance(item[3],(int,float)) else 0
            exp = item[4] if isinstance(item[4],(int,float)) else 0
            tot=round(metal+labor+exp,2)
            if tot:
                excel_cost[item[1]] = (metal,labor,exp,tot,day+' 导入')

    def creat_change():
        # 成本库：[编码，材料，人工，费用，总成本，日期，备注]
        # all_cost:[编码，材料，人工，费用，总成本，日期]

        for key, item in excel_cost.items():
            if key not in all_cost:
                change[key]=((key,) + item)
            elif key in all_cost and abs(item[3] - all_cost[key][3]) > 1:
                change[key]=((key,) + item)
                old.append(key)  # 成本已变化的，需要对数据内字段备注更改为old
                draw,name=get_code_info(key)
                change_list.append((1, key,draw,name)+item + ('导入成本变化',))
                change_list.append((2, key,draw,name)+all_cost[key] + ('原成本',))           

    rst = {}
    change = {}  #要写入数据库的数据，内部为元组
    old=[]
    change_list=[]
    excel_cost = {}

    if not ('code' and 'cost_mt' and 'cost_lb' and 'cost_exp') in excel_bom[0]:
        rst['error'] = 'BOM缺少必要的属性列（编码，成本）'
        return rst
    else:
        del excel_bom[0]    #删除BOM表里面的属性头
    creat_cost()
    creat_change()    
    
    try:
        if old:
            old = old+['code', ]
            set_old_item(tuple(old), 'cost')
        if change:
            rst['change'] = change_list
            rst['info'] = '已成功更新成本库,新增成本:%d,成本发生变化的有:%d' % (
                len(change)-len(change_list), len(change_list))
            insert_db(change.values(), 'cost')
            load_cost_db()
        else:
            rst['error'] = '成本没有发生变化'
        conn.commit()
    except Exception as ex:
        conn.rollback()
        rst['error'] = '文件写入失败: '+str(ex)
    
    return rst

def reclac_parent_cost():  # 对BOM库中组合件的成本重新按结构进行累加    
    def find_child_cost(x, n,c0,c1,c2,c3):
        if x in parent_cost and x not in updated_code:   # 只针对24R组合件或成品吗需要对子零件成本累加,其它的无须更新                
            for item in all_batch_bom[x]:
                c0,c1,c2,c3 = find_child_cost(item[0], item[1], parent_cost[x][0], parent_cost[x][1], parent_cost[x][2], parent_cost[x][3])
                
                parent_cost[x] = (
                    round(parent_cost[x][0] + c0*item[1],2),
                    round(parent_cost[x][1] + c1*item[1],2),
                    round(parent_cost[x][2] + c2*item[1],2),
                    round(parent_cost[x][3] + c3*item[1],2)
                    )

            updated_code[x] = ''
            return parent_cost[x]
        elif x in parent_cost:
            return round(parent_cost[x][0], 2), round(parent_cost[x][1], 2), round(parent_cost[x][2], 2), round(parent_cost[x][3], 2)
            
        elif x in all_cost:
            updated_code[x] = ''
            return round(all_cost[x][0] ,2),round(all_cost[x][1],2),round(all_cost[x][2],2),round(all_cost[x][3],2)
        else:
            updated_code[x] = ''
            return 0,0,0,0

    updated_code = {}
    parent_cost={}
    change = {}
    old = []
    change_list = []
    rst={}
    for f,item in all_cost.items():   #建立有成本数据的父项物料库
        if f in all_batch_bom and (f.startswith('24R') or f.startswith('28R') or f.startswith('C')):
            parent_cost[f] = (0,0,0,0)
    
    for x in parent_cost:   #对每个物料成本重新进行计算
        parent_cost[x] = find_child_cost(x, 1, 0, 0, 0, 0)

    for key, item in parent_cost.items():
        # 根据计算后的成本和原成本比较，如果发生变化则对数据库进行更改
        if item[3] and abs(item[3] - all_cost[key][3]) > 1:
            item+=(day+' 重算',)
            change[key]=((key,) + item)
            old.append(key)
            draw, name = get_code_info(key)
            change_list.append((1, key, draw, name) + item)            
            change_list.append((2, key, draw, name) + all_cost[key][:])

    try:
        if old:
            old=old+['code']
            set_old_item(tuple(old),'cost')
        if change:
            rst['change']=change_list
            insert_db(change.values(), 'cost')
            load_cost_db()
        else:
            rst['error'] = '成本没有变化'
        conn.commit()
    except Exception as ex:
        conn.rollback()
        rst['error'] = '成本更新失败:'+str(ex)    
    return rst

def scan_path(filedir,type=('.SLDDRW','.DWG')):
    def get_file(fdir):
        files = [x for x in listdir(fdir)]    #列出当前目录下所有内容
        patchs = [path.join(fdir, x) for x in files]  #拼接出当前目录下所有路径
        
        for item in patchs:
            if path.isfile(item):
                fname = path.basename(item).upper()
                lname=path.splitext(item)[1].upper()
                if lname in type:
                    name=fname.replace(lname,'')
                    file_path.append([name,item])
            elif path.isdir(item):
                get_file(item)

    file_path = []
    get_file(filedir)
    rst = {}
    if file_path:        
        rst['path']=file_path
    else:
        rst['error'] = '选择的目录内没有图纸文件,请重新选择'
    return rst

def read_json_date(filename):  # 从现有文件读取数据
    try:
        with open(filename, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        print(filename+' 数据库读取失败')
        return {}

def creat_db():  #数据库初始化检查,如果没有则创建对应的表    
    cur = conn.cursor()
    batchBOM = "CREATE TABLE IF NOT EXISTS batchBOM (\
                code VARCHAR, \
                child VARCHAR, \
                qty INTEGER NOT NULL, \
                PRIMARY KEY(code, child)); "
        
    code = "CREATE TABLE IF NOT EXISTS code (\
            code VARCHAR PRIMARY KEY,\
            draw VARCHAR,\
            name VARCHAR,\
            remark VARCHAR DEFAULT ''); "
        
    root = "CREATE TABLE IF NOT EXISTS root (\
            code VARCHAR, \
            name VARCHAR, \
            remark VARCHAR, \
            PRIMARY KEY(code, remark));"

    cost = "CREATE TABLE IF NOT EXISTS cost (\
            code VARCHAR,\
            cost_mt INTEGER,\
            cost_lb INTEGER, \
            cost_exp INTEGER, \
            cost_tot INTEGER, \
            datetime VARCHAR,\
            remark VARCHAR DEFAULT NULL); "

    # 插入默认时间:datetime VARCHAR DEFAULT (datetime('now','localtime'))

    designBOM = "CREATE TABLE IF NOT EXISTS designBOM (\
                code VARCHAR, \
                child VARCHAR,\
                qty INTEGER NOT NULL,\
                PRIMARY KEY(code, child));"

    designCODE = "CREATE TABLE IF NOT EXISTS designCODE (\
                code VARCHAR PRIMARY KEY,\
                draw VARCHAR,\
                name VARCHAR,\
                metal VARCHAR,\
                weight INTEGER,\
                remark VARCHAR); "
    
    drawPATH="CREATE TABLE IF NOT EXISTS drawPATH (\
            draw VARCHAR,\
            path VARCHAR PRIMARY KEY,\
            remark VARCHAR); "
    for sql in (batchBOM, code, root, cost, designBOM,designCODE,drawPATH):        
        try:
            cur.execute(sql)
        except Exception as ex:
            print("建表时出现如下异常:",ex)            
            
    conn.commit()
    cur.close()

def load_batch_db():  #读取数据库中小批物料和BOM信息
    cur = conn.cursor()    
    cur.execute('SELECT * FROM batchBOM')
    for item in cur.fetchall():            
        if item[0] not in all_batch_bom:
            all_batch_bom[item[0]] = [[item[1], item[2]]]
        else:
            all_batch_bom[item[0]].append([item[1],item[2]])

    cur.execute('SELECT * FROM code')
    for item in cur.fetchall():
        all_batch_code[item[0]] = (item[0] + item[-1], item[1], item[2])

    cur.close()

def load_root_db():  #读取已写入的root列表,分为batchBOM和designBOM    
    cur = conn.cursor()
    cur.execute('SELECT * FROM root')
    for item in cur.fetchall():
        if item[2]=='BATCH':
            batch_root[item[0]] = item[1]
        elif item[2] == 'DESIGN':
            design_root[item[0]] = item[1]
        elif item[2] == 'PATH':
            path_root[item[0]]=item[1]
    cur.close()

def load_cost_db():  #读取成本信息    
    cur = conn.cursor()    
    cur.execute('SELECT * FROM cost')
    for item in cur.fetchall():   #cost库：[0编码，1材料，2人工，3费用，4总，5日期，6备注]
        if not item[-1]:
            all_cost[item[0]] = item[1:6]
        
    cur.close()

def load_design_db():  #读取设计BOM  
    cur = conn.cursor()
    cur.execute('SELECT * FROM designBOM')
    for item in cur.fetchall():
        if item[0] not in all_design_bom:
            all_design_bom[item[0]] = [(item[1], item[2])]
        else:
            all_design_bom[item[0]].append((item[1], item[2]))
        
    for item in all_design_bom.values():
    #考虑到物料结构借用，所以要将设计BOM中用到的组合件(24R/28R)，但又没有子零件的，从小批库读取过来
        for m in item:
            if m[0][:3] in ('24R', '28R') and m not in all_design_bom and m in all_batch_bom:
                all_design_bom[m]=all_batch_bom[m]

    cur.execute('SELECT * FROM designCODE')
    for item in cur.fetchall():
        all_design_code[item[0]] = item
    cur.close()

def load_drawpath_db():
    cur = conn.cursor()
    cur.execute('SELECT * FROM drawPATH')
    for item in cur.fetchall():
        draw_path[item[0]] = item[1]

def insert_db(new,sheet):   #将信息写入数据库    
    cur = conn.cursor()
    table = {
        'batchBOM': ' VALUES (?,?,?)',
        'code': ' (code,draw,name) VALUES (?,?,?)',
        'root': ' VALUES (?,?,?)',
        'cost': ' (code,cost_mt,cost_lb,cost_exp,cost_tot,datetime) VALUES (?,?,?,?,?,?)',
        'designBOM': ' VALUES (?,?,?)',
        'designCODE': ' VALUES (?,?,?,?,?,?)',        
        'drawPATH': ' VALUES (?,?,?)',
    }

    sql = 'INSERT OR REPLACE INTO ' + sheet + table[sheet]
    cur.executemany(sql, new)
    cur.close()

def remove_db(old1='',col1='code',old2='',col2='',sheet=''):   #从数据库中删除指定编码的物料    
    cur = conn.cursor()
    sql = 'DELETE FROM ' + sheet + ' WHERE ' + col1 + ' in ' + str(old1)
    if col2 and old2:
        sql += 'AND ' + col2 + ' in ' + str(old2)
        
    cur.execute(sql)
    cur.close()

def set_old_item(olditem,table):  #对有变动的物料,remark设置为old    
    cur = conn.cursor()
    sql = 'UPDATE '+table+' SET remark =\'old\' WHERE code in ' + str(olditem)    
    cur.execute(sql)    
    cur.close()

def find_db(col,item, sheet):   #直接从数据库查询物料    
    cur = conn.cursor()
    
    if item == '*':
        sql = 'SELECT * FROM ' + sheet
    else:
        sql = "PRAGMA table_info({0})".format(sheet)
        cur.execute(sql)
        cols = tuple(x[1] for x in cur.fetchall())  #得到数据表所有列名
        if col == 'ALL':
            item = '\'%#' + item + '#%\''
            col=''
            for item in cols:
                if col:
                    col+='\'#\'||'
                col += item
            col += '\'#\''
            col = '\'#\''+col
            
            #col = '\'#\'code\'#\'||draw\'#\'||name\'#\''  # 将各个字段拼接起来，用like进行比较
        elif col in cols:
            item = '\''+item+'\''
        else:       
            return
    
        sql = 'SELECT * FROM ' + sheet + ' WHERE ' + col + ' LIKE ' + item

    cur.execute(sql)
    rst=cur.fetchall()
    cur.close()
    return rst

def load_old_cost():  #从数据库查询有变动过的成本
    cost_changed = []    
    code=''
    cur = conn.cursor()
    cur.execute('SELECT * FROM cost WHERE remark=\'old\' ORDER BY code ASC')    
        
    for item in cur.fetchall():  # cost库：[0编码，1材料，2人工，3费用，4总，5日期，6备注]
        if item[0] != code:
            code=item[0]
            draw, name = get_code_info(code)
            cost_changed.append((1,code,draw,name)+ all_cost[item[0]])
            cost_changed.append((2, code,draw,name) + item[1:])
        elif item[0] == code:
            draw, name = get_code_info(code)
            cost_changed.append((2, code,draw,name) + item[1:])    
    cur.close()
    return cost_changed

def save_to_excel(file, bom,target,colname):  #把表格内容保存到excel文件

    wb = xl.Workbook()
    ws = wb.active
    fill_blue = PatternFill('solid',fgColor='00B2EE') #设置填充颜色为 橙色 
    font_title = Font(u'微软雅黑', size=14, bold=True, italic=True)  #设置字体样式
    font_bold = Font(u'微软雅黑', bold=True,)
    
    ws.append([target])     #添加标题    
    ws['A1'].font = font_title

    ws.append(colname)   #添加列名
    for i in ws[2]:
        i.fill = fill_blue
        i.font = font_bold
    
    lv_n = [0,0,0,0,0,0,0,0,0]       
    for key in bom:
        lv_n[key[0] + 1:8] = 0,0,0,0,0,0,0,0
        lv_n[key[0]] += 1
        
        key=[x if x!='-' else '' for x in key]
        ws.append(['+'*key[0],lv_n[key[0]]]+key[1:])

    ws.row_dimensions[1].height = 35
    ws.row_dimensions[2].height = 20 
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 15
    rst={}
    try:
        wb.save(file)        
    except Exception as ex:
        rst['error']='文件保存错误: '+str(ex)
    return rst
    
#---------------------主程序区---------------------

all_batch_bom = {}
all_batch_code = {}
all_cost = {}
all_design_bom = {}
all_design_code={}
design_root = {}
batch_root = {}
path_root = {}
draw_path={}
day = datetime.now().strftime('%Y-%m-%d')
#day='2019-10-23'

dbfile = 'batchITEM.db'
conn = sqlite3.connect(dbfile)

rule = {'code': r'\d{2}R',
        'root': r'C|E\d{2}-',
        'asmb': r'24|8R',
        'metal': r'0[1234]R',
        'rootnew':r'N\d{4}'
            }

creat_db()
load_root_db()
load_cost_db()
load_batch_db()
load_design_db()
#load_drawpath_db()

print('code库记录: ', len(all_batch_code))
print('cost库记录:',len(all_cost))
print('batchBOM库记录: ' ,len(all_batch_bom))
print('designBOM库记录：',len(all_design_bom))
print('小批产品库记录：', batch_root)
print('设计产品库记录：', design_root)
#print('图纸库记录：', len(draw_path))

op = main_GUI()
op.root.mainloop()
