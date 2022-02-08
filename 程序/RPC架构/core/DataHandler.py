# -*- coding:utf-8 -*-
# 数据处理函数

from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import json, sqlite3, re,logging,traceback
from os import path,listdir,walk
from datetime import datetime

class dataHandler():
    def __init__(self):
        self.today = datetime.now().strftime('%Y-%m-%d')
        self.conn = sqlite3.connect('batchITEM.db')
        # 编码的匹配规则
        #all_cost={}
        self.rule = {'code': r'\d{2}R\d{7}',
                'batch': r'C|E\d{2}-|D',
                'asmb': r'24|8R',
                'metal': r'0[1234]R',
                'root': r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})',
                'design': r'N\d{4}',
                'custom': r'YF\d{7}',
                'exper': r'(EXP|CS)\d{7}',
                'temp': r'TEMP\d{2}'
                }
        self.all_root = {'BATCH':{},'DESIGN':{},'CHANGE':{},'PATH':{}}
        self.all_code={}
        self.all_batch_bom={}
        self.all_design_code={}
        self.all_design_bom=[]
        self.design_change_bom=[]
        self.draw_path={}
        self.all_cost={}
        self.old_cost=[]
        try:
            self.logger()
            self.creat_db()
            self.all_code, self.all_batch_bom,self.all_root['BATCH'] = self.load_batch_db()
            self.all_design_code = self.load_designCODE()
            self.all_code.update(self.all_design_code)    #将设计物料和小批物料合并在一起

            self.all_design_bom,self.all_root['DESIGN']=self.load_designBOM()
            self.design_change_bom,self.all_root['CHANGE']=self.load_design_change()
            self.draw_path,self.all_root['PATH']=self.load_drawpath_db()
            self.all_cost, self.old_cost=self.load_cost_db()
            #self.all_root['BATCH'] = self.get_batchRoot()
            #self.all_root['DESIGN'] = self.get_designRoot()
            #self.all_root['CHANGE'] = self.get_designChangeRoot()
            #self.all_root['PATH'] = self.get_drawPathRoot()
            
        except Exception as e:
            logging.warning(e, exc_info=True, stack_info=True)
            print(e)
        else:
            info = "code {0},batchBOM {1},designBOM {2}".format(
                len(self.all_code), len(self.all_batch_bom), len(self.all_design_bom),)
            print(info)
            logging.info('LOAD DB... '+info)

    def logger(self,):
        '程序LOG记录'
        LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
        DATE_FORMAT = "%Y-%m-%d %H:%M:%S"

        logging.basicConfig(filename='log.log', level=logging.DEBUG,
                            format=LOG_FORMAT, datefmt=DATE_FORMAT)

    def str_to_var(self,x):
        if isinstance(x, str):
            if x.upper()=='BATCH':
                return self.all_batch_bom
            elif x.upper()=="DESIGN":
                return self.all_design_bom
            elif x.upper()=="DESIGNCHANGE":
                return self.design_change_bom
            else:
                return x
        else:
            return x

    def find_code(self,f,db={},opt='AND'):
        '根据输入内容在小批和设计物料库查找物料，f为查找的字段列表，exact为TRUE则准确匹配，当字符串中包含空格时，会被分割并分别查找'
        rst_code = []
        rst={}
        if not isinstance(f,(list,tuple)):
            rst['error']='输入的不是列表'
            return rst

        for item in self.all_code.values():
            tar=False
            for x in f:
                if x in str(item):
                    tar=True
                    if opt=='OR':
                        break
                else:
                    tar=False
                    if opt=='AND':
                        break
            if tar:
                rst_code.append(item[:])

        if rst_code:
            for n,item in enumerate(rst_code):
                if item[0] == item[1] + item[2]:    #对于设计BOM中编码=图号+名称的,编码用''代替        
                    rst_code[n] = ('',) + item[1:]

            rst['code'] = rst_code  #[编码，图号，名称,时间].
        
        return rst

    def check_code(self,code, draw,name):
        '对物料编码进行确认：如果输入了编码，则根据末尾和图纸查找是不是最新版；如果没有输入编码，则先根据图号查找有无唯一编码，若无则根据名称查找有无唯一编码。如果查找到的是同一个编码不同版本，则返回最新编码'
        def rst_check(rst):
            new_code=''
            codes = []
            if 'code' in rst:
                for i in rst['code']:
                    if re.match(self.rule['code'],i[0]):  #挑选出有真正编码的，去掉设计物料
                        codes.append(i[0])
            if codes:
                codes.sort()
                codes.reverse()
                if len(codes)==1:
                    new_code=codes[0]
                elif codes.count(codes[0][:-1])==len(codes):  #当查询的编码除最后一位都相同时
                    new_code=codes[0]
            
            return new_code

        new_code = ''        
        rst={}
        rst_={}
        rst['old']=code
        if draw != '' and 'GB' not in str(draw):
            rst_ = self.find_code((draw,), db='BATCH')
            new_code=rst_check(rst_)
        if not new_code:        
            rst_ = self.find_code((name,), db='BATCH')
            new_code = rst_check(rst_)

        if not new_code or new_code==code:   #当没找到新的编码时或者查找结果和原来相同，返回旧编码
            pass
        elif not code or code == ' ':   #当原编码为空时，直接写入
            rst['new']=new_code
            rst['style']='green'
        else:     #当原编码存在时，进行对比
            if code[:-1]==new_code[:-1]:   #当找到的新旧编码除最后一位外相同时
                if code[-1]<new_code[-1]:
                    rst['new'] = new_code
                    rst['style']='blue'
                    rst['remark']='原编码为老版本：'+str(code)                
                else:
                    pass
            else:
                rst['new'] = new_code
                rst['style']='red'
                rst['remark']='由图号查询的编码不同，原内容:\n'+str(code)
        return rst

    def get_code_info(self,code):
        '根据编码返回对应的编码、图号、名称、材料、重量、备注' 

        if re.match(self.rule['code'],code) and code[-1]=='P':
            code=code[:-1]

        if code in self.all_code:
            item=self.all_code[code][:-1]
        else:
            item=(code,"","","","","",)

        return item

    def find_parent_bom(self,f,all_bom):
        "根据编码反查使用的BOM,在字典每个值的里面查找编码,找到后将对应的key,再作为编码进行同样查找"
        
        def find_parent(x, n):
            parent[n]=''
            for key, item in all_bom.items():
                for code in item:
                    if x in code:
                        parent[n] = [n] + list(code)
                        find_parent(key, n - 1)
                        break
                    
            if not parent[n] and x!=f:
                parent[n] = [n] + [x, 1]
                if x not in father_bom_dict:
                    father_bom_dict[x] = []
                num = n - 1
                p1=[]
                for i in range(n, 10):                               
                    p1.append([parent[i][0] - num, parent[i][1], parent[i][2]])
                father_bom_dict[x].append(p1[:])

        def fmt_father_bom():    # 去除相同的父项,并由字典转化为列表形式
            for key in father_bom_dict:
                lvcode = {}
                father_bom_dict[key].sort()
                for items in father_bom_dict[key]:
                    for item in items:
                        if lvcode.get(item[0], 'NA') != item[1]:
                            lvcode[item[0]] = item[1]
                            for n in range(item[0] + 1, 7):
                                lvcode[n] = ''
                            bom.append([item[0], item[1], item[2]])

        def father_bom_total(x):
            # 计算反查物料在顶层的总用量,和本层用量
            lv_num = {0: 1, 1: 1}
            code_num = [0]
            root_index = 0

            for n, item in enumerate(bom):
                lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
                item.append(lv_num[item[0]])

                if item[1] == x:
                    code_num.append(lv_num[item[0]])

                if item[0] == 1 or n == len(bom) - 1:
                    bom[root_index][3] = (sum(code_num))
                    code_num = [0]
                    lv_num = {0: 1, 1: 1}
                    root_index = n

        bom = []
        parent={}
        father_bom_dict = {}
        rst={}

        find_parent(f, 9)    
        if father_bom_dict:
            fmt_father_bom()
            father_bom_total(f)
            for n,item in enumerate(bom):
                code,draw,name,metal,weight,remark=self.get_code_info(item[1])
                if code == draw + name:
                    code = ''
                bom[n] = (item[0], code, draw, name, item[2], item[3], metal, weight, remark)

            rst['bom']=bom
        return rst

    def find_child_bom(self,f,all_bom):
        "根据编码查找BOM中的子零件，分小批和设计BOM"
        def bom_total():
            # 计算反查物料在顶层的总用量,和本层用量
            lv_num = {0: 1, 1: 1}
            for item in bom:
                lv_num[item[0]] = item[2] * lv_num[item[0] - 1]
                item.append(lv_num[item[0]])
        
        def find_child(x, n):
            if x in all_bom:
                for item in all_bom[x]:
                    bom.append([n]+list(item))
                    find_child(item[0],n+1)
        bom = []
        rst = {}
        all_bom=self.str_to_var(all_bom)
        if f in all_bom:
            bom.append([1,f,1])
            find_child(f, 2)
            bom_total()
            for n,item in enumerate(bom):
                code,draw,name,metal,weight,remark=self.get_code_info(item[1])
                if code == draw + name:
                    code = ''
                bom[n]=(item[0],code,draw,name,item[2],item[3],metal,weight,remark)

            rst['bom'] = bom
        return rst

    def update_to_bom_db(self,excel_bom, db='BATCH', type='R'):
        '将已读取的BOM写入BOM库,先检查是否有新物料，有编码的写入小批物料库，无编码的写入设计物料库；同时将设计物料属性写入小批物料库中。再检查有无新BOM结构，有则进行写入'
        #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注]
        # 把设计BOM按小批样式，分物料库和BOM库

        def creat_new_designcode():  # 当读取的是设计BOM时，无编码的物料保存到designCODE；有编码的属性添加到小批库
            for item in excel_bom:

                if item[1] in self.all_code:   #对于有编码的，检查小批物料库属性，不同则添加
                    if tuple(item[5:8])!=self.all_code[item[1]][3:6]:
                        new_batchcode[item[1]]=self.all_code[item[1]][:3]+(item[5],item[6],item[7],self.all_code[item[1]][-1],)
                else:
                    new_batchcode[item[1]]=(item[1],item[2],item[3],item[5],item[6],item[7],self.today)

                if item[1] == '':  # 对于找不到编码的，用图号+名称作为编码
                    item[1] = item[2] + item[3]
                    if type=='W' or item[1] not in self.all_code:
                        new_designcode[item[1]] = (
                            item[1], item[2], item[3], item[5], item[6], item[7], self.today)

        def creat_designbom():
            sn={}
            order=[0,0,0,0,0,0,0,0,0,0]
            for item in excel_bom:
                order[item[0]+1:]=0,0,0,0,0,0,0,0,0
                order[item[0]]+=1

                if item[0]==1:
                    rt=item[1]
                    lv='ROOT'
                    sn[1]='1'
                else:
                    lv=sn[item[0]-1]+'.'+str(order[item[0]])
                    sn[item[0]]=lv
                    
                new_bom.append((lv,item[1],item[4],rt))        

        def creat_bom_dict():  #创建读取物料的设计BOM的字典
            bom={}
            lv_code = {}
            ex_code_lv = 10
            
            for n in excel_bom:
                code = n[1]
                lv = n[0]
                num = n[4]
                if lv > ex_code_lv:
                    continue
                elif lv>1:
                    bom[lv_code[lv - 1]].append((code, num))

                if code in bom:
                    ex_code_lv = lv
                else:
                    lv_code[lv] = code
                    bom[code] = []
                    ex_code_lv = 10        

            for key, item in bom.items():
                if item:
                    ebom_dict[key] = item

        def creat_new_bom(): 
            #将读取的设计BOM和已有的设计BOM结构进行对比，新的添加到new_bom，有变动的物料添加到old，然后到designBOM删除结构
            for key, item in ebom_dict.items():
                if key not in all_bom:
                    for code in item:
                        new_bom.append((key, code[0], code[1]))
                elif type=='W': #因为内部元素顺序可能不同，所以要转为set，再进行对比                
                    if set(item) != set(all_bom[key]):
                        old_bom.append(key)
                        for code in item:
                            new_bom.append((key, code[0], code[1]))

        if not ('lv' and 'code' and 'name' and 'num') in excel_bom[0]:
            return 'BOM缺少必要的属性列（层次，编码，名称，数量）'
        else:
            del excel_bom[0]  #删除BOM表里面的属性头

        new_batchcode={}
        ebom_dict={}
        #new_code = {}   #新的物料,用于写入小批物料库
        new_designcode={}   ##新的物料,用于写入设计物料库    
        new_bom = []    #新的BOM层次,用于写入数据库
        old_bom = []        #有变动的物料,用于在数据库中删除
        
        #creat_bom()
        if db == 'BATCH':
            all_bom = self.all_batch_bom
            dbsheet_bom = 'batchBOM'
            creat_bom_dict()
            creat_new_bom()
        elif db in ('DESIGN','CUSTOM','EXPER'):
            #all_bom = all_design_bom
            dbsheet_bom = 'designBOM'       
            creat_new_designcode()
            creat_designbom()

        root = excel_bom[0][1]
        rootname = excel_bom[0][3]
        
        try:
            if new_designcode:
                self.insert_db(list(new_designcode.values()), 'designCODE')            
                self.all_code.update(new_designcode)
            
            if new_batchcode:
                self.insert_db(list(new_batchcode.values()), 'batchCODE')

            #if re.match(rule['root'], root):
            #    insert_db(((root, rootname, db),), 'root')  #将root信息写入root
                
            if old_bom:
                self.remove_db(tuple(old_bom), sheet=dbsheet_bom)
            if new_bom:
                self.insert_db(new_bom, dbsheet_bom)        
            self.conn.commit()
            return '已成功写入'
        except Exception as ex:
            self.conn.rollback()
            logging.debug(str(traceback.format_exc()))
            return '库写入失败,已撤销:'+ str(ex)

    def update_to_code_db(self,excel_bom):
        '根据读取的物料信息,将变动过的写入数据库，同时检查新旧编码，并对旧的标记old'    
        #[0编码,1图号,2名称,3材料,4重量,5备注,6时间]

        def creat_code_dict(): #创建读取物料的字典库，同时去除了重复项
            for item in excel_bom:
                excel_code[item[0]] = tuple(item)
            if 'code' in excel_code:
                del excel_code['code']
                
        def creat_new(): #和原来物料库进行对比，找出新增和修改的物料
            for key, item in excel_code.items():            
                if key not in self.all_code:
                    change.append(item)
                elif item[1] != self.all_code[key][1] or item[2] != self.all_code[key][2].replace('old', ''):
                    change.append(item)
                    mod_list.append((1,'新修改:'+item[0],item[1],item[2]))
                    mod_list.append((2,'  原物料:'+item[0],self.all_code[key][1],self.all_code[key][2]))

        rst = {}
        excel_code={}   #读取物料的字典
        change = []    #要写入数据库的物料
        #old=[]  #核对的旧编码，加old标记
        mod_list=[] #修改过的物料列表
        
        if not ('code' and 'name') in excel_bom[0]:
            return '编码表缺少必要的属性列（编码,名称）'
        else:
            del excel_bom[0]  #删除BOM表里面的属性头

        creat_code_dict()
        creat_new()
        self.all_code.update(excel_code)
        #old_item_check()
        
        if mod_list:
            rst['mod'] = mod_list
        try:
            #if old:
            #    set_old_item(tuple(old), 'code')
            if change:
                rst['new'] = len(change)-len(mod_list)/2
                self.insert_db(change, 'batchCODE')
            self.conn.commit()
        except Exception as ex:
            self.conn.rollback()
            rst['error']='写入文件失败:'+str(ex)
            logging.debug(str(traceback.format_exc()))

        return rst

    def update_to_cost_db(self,excel_bom,day):
        '根据读取的BOM,找出成本变动过的,写入数据库;对原来的成本标记old'
        #[0编码,1名称,2材料成本,3人工成本,4管理成本,5导入日期,6标记] 
        def creat_cost():
            for item in excel_bom:          
                tot=round(item[2]+item[3]+item[4],2)
                if tot:
                    excel_cost[item[0]] = (item[2],item[3],item[4],tot,day)

        def creat_change():
            # 成本库：[编码，材料，人工，费用，总成本，日期，备注]
            # all_cost:[编码，材料，人工，费用，总成本，日期]

            for key, item in excel_cost.items():
                if key not in all_cost:
                    change[key]=((key,) + item+('',))
                elif key in all_cost and abs(item[3] - all_cost[key][3]) > 1:
                    if day < all_cost[key][4]:    #导入日期和原成本日期对比，决定哪个成本作为最新成本
                        change[key]=((key,) + item+('OLD',))
                    else:
                        change[key]=((key,) + item+('',))
                        old.append(key)     # 成本已变化的，需要对数据内字段备注更改为old
                    cd,draw,name=self.get_code_info(key)[:3]
                    change_list.append((1, key,draw,name)+item + ('导入成本变化',))
                    change_list.append((2, key,draw,name)+all_cost[key] + ('原成本',))           

        rst = {}
        change = {}  #要写入数据库的数据，内部为元组
        old=[]
        change_list=[]
        excel_cost = {}

        if not ('code' and 'cost_mt' and 'cost_lb' and 'cost_exp') in excel_bom[0]:
            rst['error'] = 'BOM缺少必要的属性列（编码，成本）'
            return rst
        else:
            del excel_bom[0]    #删除BOM表里面的属性头

        
        all_cost,old_cost=self.load_cost_db()
        creat_cost()
        creat_change()    
        
        try:
            if old:
                old = old+['code', ]
                self.set_old_item(tuple(old), 'cost')
            if change:            
                rst['info'] = '已成功更新成本库,新增成本:%d,成本发生变化的有:%d' % (
                    len(change)-len(change_list), len(change_list))
                self.insert_db(list(change.values()), 'cost')            
            else:
                rst['info'] = '成本没有发生变化'
            self.conn.commit()
        except Exception as ex:
            self.conn.rollback()
            rst['error'] = '文件写入失败: ' + str(ex)
            logging.debug(str(traceback.format_exc()))
            
        if change_list:
            rst['change'] = change_list
        return rst

    def update_to_change_db(self,excel_bom):
        def creat_change_item():
            '将读取的设计更改内容，添加文件名，并检查序号是否重复   (设计更改文件名+序号)'
            name=''        
            for row,item in enumerate(excel_bom[1:]):
                if item[0]=='ROOT':
                    name=item[2]
                if not name:
                    return '第{0}行缺少更改文件名'.format(str(row+1))

                if not item[0]:
                    return '第{0}行缺少序号'.format(str(row+1))

                if name and item[0]:
                    s1=name+'#'+item[0]
                    if s1 not in change_item:
                        change_item[s1]=tuple(item+[name])

                    else:
                        return '第{0}行序号重复'.format(str(row+1))

        change_item={}
        rst={}    
        err=creat_change_item()
        if err:
            rst['error']=err
            return rst

        try:
            if change_item:
                self.insert_db(list(change_item.values()), 'design_change_item')
            self.conn.commit()        
            self.load_design_change()
        except Exception as ex:
            self.conn.rollback()
            rst['error']='写入数据库失败:'+str(ex)
            logging.debug(str(traceback.format_exc()))   
        return rst

    def update_to_drawpath_db(self,name,pathdir,bom):
        rst={}
        num=len(bom)
        
        bom=[tuple(x+[name,]) for x in bom]
        bom.insert(0,('ROOT',pathdir,self.today,name))
        try:
            #insert_db(((name, pathdir, 'PATH'),), 'root')
            self.insert_db(bom, 'drawPATH')
            self.conn.commit()
            rst['txt']=str(name)+' 已更新%d 记录: '%num        
        except Exception as ex:
            self.conn.rollback()
            rst['txt']=str(ex)+' 更新失败'
        
        return rst

    def find_in_bom(self,xx,bom,parent=False,child=False,opt='AND'):
        '在BOM层次中查找物料: xx是要查找的字段列表;     bom是被查找的列表,如果输入的是字符串，则转换为对应的bom;    parent、child参数控制是否附加父项或子项;    opt参数决定用AND还是OR,按正向查找'
        rst={}
        bom = self.str_to_var(bom)
        bom=list(bom)
        new_bom=[]
        new_item=[]
        if isinstance(xx,str):
            xx=(xx,)

        lv_item={}
        k=9
        for item in bom: 
            i=item[0]
            lv_item[i]=item
            tar=False

            if child:   #控制是否添加查找内容的子零件
                if item[0] > k:
                    new_bom.append(item)
                    continue
                else:
                    k=9
            for x in xx:
                if x in str(item):
                    tar=True
                    if opt=='OR':
                        break
                else:
                    tar=False
                    if opt=='AND':
                        break
            if tar:
                k=item[0]
                new_item.append(item)
                if parent:    #控制是否添加父项
                    for key,val in lv_item.items():
                        if key <=i and val:
                            new_bom.append(val)
                            lv_item[key]=[]
                else:
                    new_bom.append(item)

        if new_bom:       
            for n,item in enumerate(new_bom):
                item=list(item)
                if item[1] == item[2] + item[3]:    #对于设计BOM中编码=图号+名称的,编码用''代替  
                    item[1]=''
                new_bom[n]=tuple(item)

            rst['num']=len(new_item)
            rst['bom']=new_bom
            rst['code'] = new_item  #[编码，图号，名称,时间].    

            sum=0  #用量
            for item in new_item:
                if len(item)>5 and isinstance(item[5],(int,float)):
                    sum+=item[5]
                else:
                    sum=0
                    break
            if sum:
                rst['sum']=sum
            
        return rst

    def remove_assemble(self,bom):
        '去除BOM中的虚拟件，有编码的则按24/28R去除，无编码的按材料属性去除；对于焊接件，则去除其子零件'
        bom_dict = {}
        single_bom = []
        del_bom=[]

        del_lv=9
        for n,item in enumerate(bom[:]):

            item = list(item[:-1])
            if item[0] > del_lv:  #删除焊接件子零件
                del_bom.append([n+1]+item)
                continue
            else:
                del_lv = 9
            #先用编码判断，再用属性判断。同时再判断是否有子零件，避免误删        
            if re.match(r'(01)|(02)|(03)|(04)|(24)|(28)R\d*',item[1]):
                del_bom.append([n+1]+item)
                continue
            elif re.search(r'(组合件)|(组装件)|(装配件)|(明细)|(明细表)|(原理)',str(item)):            
                if bom[n+1][0]<=item[0]:
                    del_bom.append([n+1]+item)
                    continue

            if re.search(r'焊接件',str(item)):   #
                del_lv = item[0]
            
            if not item[1]:
                item[1] = item[2] + item[3]

            if item[1] in bom_dict:
                bom_dict[item[1]][5] += item[5]    #对同编码的数量进行累加
            else:
                bom_dict[item[1]] = item[:]

        #bom_=sorted(bom_dict.keys(),key=lambda x:bom_order.index(x))
        bom_key=bom_dict.keys()
        #single_bom.append(bom[0][:-1])
        for item in bom_dict.values():
            if item[1]==item[2]+item[3]:
                item[1]=''
            item[0]=1       #把所有层次改为1
            single_bom.append(tuple(item))
        
        return single_bom,del_bom

    def read_design_BOM(self,file, type='BATCH'):
        '统一读取各种excel文件,根据不同类型生成不同列；读取时进行序号、编码、数量格式检查；针对check和计算数量模式会写入到原excel'
        # [0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]

        def get_col(wsheet):
            col_={}
            lable = {
                '级别': 'lv',
                '层次': 'lv',
                '序号': 'lv',
                '编码': 'code',
                '子件编码': 'code',
                '新编码':'code',
                '图号': 'draw',
                '代号': 'draw',                      
                '名称': 'name',
                '子件名称': 'name',
                '描述':'name',
                '用量': 'num',
                '数量': 'num',
                '基本用量': 'num',
                '使用数量' :'num',           
                '材料': 'metal',
                '子件规格': 'metal',
                '单重':'weight',
                '备注': 'remark',
                '材料成本': 'cost_mt',
                '人工成本': 'cost_lb',     
                '费用成本': 'cost_exp',
                '日期':'time',
                '序号1':'sn',
                '零部件图号':'draw',
                '零部件名称':'name',
                '更改前编码':'code_old',
                '更改后编码':'code_new',
                '更改前说明':'notice',
                '更改后说明':'notice_new',
                '更改类别':'change_lv',
                '更改方式':'change_draw',
                '库存':'stock',
                '在途':'on_order',
                '已制品处理建议':'stock_note',
                '涉及机型':'related_model',
                }

            lable_t={}
            for key,item in lable.items():
                if item in title:
                    lable_t[key] = item
                    
            str_row = 1
            for row in wsheet.values:  #在工作表逐行查找表头
                col_['used']=[]
                str_row += 1
                if str_row > 10:
                    break
                for c, value in enumerate(row):
                    for key in lable_t:
                        if isinstance(value, str) and key == value.replace(' ', ''):
                            col_['used'].append(c)
                            if lable_t[key] == 'lv':
                                if 'lv' not in col_:
                                    col_['lv']=[c,]
                                else:
                                    col_['lv'].append(c)
                            elif lable_t[key] in col_:
                                return '%s 的属性列重复'% key
                            else:
                                col_[lable_t[key]] = c                        

                if 'name' in col_ and len(col_)>=2:                
                    if 'lv' in title and 'lv' not in col_:
                        return '找不到层次列'
                    if 'num' in title and 'num' not in col_:
                        return '找不到数量列'
                    if 'time' in title and 'draw' in col_:
                        col_['time']=col_['draw']+1

                    col_['str_row'] = str_row
                    return col_
                else:
                    col_.clear()

            return '找不到基本的属性列:' + str(title)

        def read_item(wsheet):
            def fmt_str(x):
                if x:
                    x=str(x)
                    x = x.replace(' ', '')
                    if x:
                        return x.upper()
                    else:
                        return ''
                else:
                    return ''
            def fmt_num(x):
                if not x or x==' ':
                    return 0
                elif isinstance(x, int):
                    return x            
                else:
                    try:                
                        return round(float(x), 2)
                    except:
                        return x
            def code_rule(x):
                if re.match(self.rule['code'], x):
                    return True
                elif re.match(self.rule['root'], x):
                    return True            
                elif re.match(self.rule['temp'], x):
                    return True
                else:
                    return False
            def fmt_time(x):           
                if x:
                    x = str(x)
                    x = x.replace(' ', '')
                else:
                    return ""

                t = r'(\d+)[\.\-\\\/](\d+)[\.\-\\\/](\d\d?)'
                tt = re.findall(t, x)
                if tt and len(tt[0])==3:
                    yy,mm,dd = tt[0]
                
                    if len(yy) == 2:
                        yy = '20' + yy
                    if len(mm) == 1:
                        mm = '0' + mm
                    if len(dd) == 1:
                        dd = '0' + dd
                    elif len(dd) > 2:
                        dd=dd[:2]
                    return yy + "-" + mm + "-" + dd
                else:
                    return ""
            def get_lv(lvs):
                #if not excel_bom:   #先要找到ROOT才能继续
                #    if isinstance(lvs[0], str) and lvs[0].upper() == 'ROOT':                
                #        return 1
                #    else:                
                #        return ' 行没有ROOT'
                if len(lvs)==1:    #针对层次只有1列：一种是'+++'，另一种是数字加.来区分
                    if not lvs[0]:
                        return ' 缺少层次'
                    lv1=str(lvs[0])
                    if '+' in lv1:
                        lv = len(lv1)
                    else:
                        lv=lv1.count('.')+2
                    return lv
                else:   # 层次由多列，EBOM格式，数字所在的列代表层次高低
                    lv_1 = []
                    for col, n in enumerate(lvs):
                        if isinstance(n, int):
                            lv = col+1
                            lv_1.append(lv)
                    if len(lv_1) == 1:
                        return lv
                    elif len(lv_1) > 1:
                        return ' 行层次重复'
                    else:
                        return ' 行缺少层次'

            row_num = 0
            lv_2 = 2
            last_lv=0        
            for row in wsheet.iter_rows(values_only=True):
                row_num += 1
                org_bom.append(list(row))
                if row_num<col['str_row']:
                    continue
                item={}
                for key in title:
                    if key in col:
                        if key == 'lv':
                            if row_num == col['str_row']:  #第一行检查有没有ROOT                            
                                if row[0] and row[0].upper() in ('ROOT', 'root'):
                                    item['lv'] = 1                                
                                else:
                                    item['lv'] =' 行没有ROOT'
                            else:
                                item[key] = get_lv([row[x] for x in col['lv']])
                        elif key in ('num','weight','cost_mt', 'cost_lb', 'cost_exp'):
                            item[key] = fmt_num(row[col[key]])
                        elif key in ('time',):
                            item[key]=fmt_time(row[col[key]])
                        else:
                            item[key] = fmt_str(row[col[key]])
                    elif key=='row':
                        item['row']=row_num
                    elif key=='other':   #属性列没用到的都添加到该项
                        item['other']=[]
                        for i,v in enumerate(row):
                            if i not in col['used']:
                                item['other'].append(v)
                    elif key=='all':
                        item['all']=list(row)
                    else:
                        item[key] = ''

                if 'name' in item and item['name'] == '':  #跳过空行
                    continue

                if 'lv' in item:
                    if isinstance(item['lv'], int):
                        if len(excel_bom)==1:   # 第二行的层次必须是2, 所以要根据第二行的层次来确定一个整体层次的调整系数
                            lv_2=item['lv']
                            last_lv=1
                            
                        if lv_2<=2:
                            item['lv']=item['lv']+2-lv_2
                        else:
                            item['lv']=item['lv']-lv_2+2

                        if item['lv'] > last_lv + 1:  # 检查层次是否连续
                            item_error.append('第 ' + str(row_num) + ' 行层次和上层脱节')
                        else:
                            last_lv = item['lv']
                    else:
                        item_error.append('第 ' + str(row_num) + item['lv'])
                                                            
                if 'num' in item:
                    if not isinstance(item['num'], (int, float)):
                        item_error.append('第 ' + str(row_num) + '行没有数量或格式不对')

                if 'cost_mt' in item:
                    try:
                        float(item['cost_mt']+item['cost_lb']+item['cost_exp'])
                    except:
                        item_error.append('第 ' + str(row_num) + '行成本不是数字格式')

                if 'code' in item:  #对编码格式进行检查,并查找对应的编码
                    if type in ('CHECK','QTY'):
                        pass
                    elif not code_rule(item['code']):
                        if item['code'] == '' and type in ('DESIGN', 'EXPER','CUSTOM',):  #允许编码为空，并记录本行信息
                            pass
                        elif type in ('CODE', 'TEMP','REMOVE', 'COST',):  #不处理
                            pass
                        else:
                            item_error.append('第 ' + str(row_num) + '行编码格式不对')

                if 'sn' in item:
                    if not item['sn']:
                        item_error.append('第 ' + str(row_num) + '行没有序号')

                if not item_error:
                    t=[]
                    for k in title:
                        if k=='other':
                            t=t+item[k]
                        else:
                            t.append(item[k])
                    excel_bom.append(t)
        
        excel_bom = []
        org_bom=[]
        #col = {}
        rst={}
        skip_sheet = ''
        item_error=[]

        if type in ('BATCH','DESIGN','EXPER','CUSTOM'):
            title = ('lv', 'code', 'draw', 'name', 'num', 'metal', 'weight', 'remark')        
        elif type == 'CODE':
            title = ('code', 'draw', 'name', 'metal', 'weight', 'remark','time')
        elif type == 'COST':
            title = ('code', 'name', 'cost_mt', 'cost_lb', 'cost_exp')
        elif type=='DESIGN_CHANGE':
            title=('sn','draw','name','code_old','code_new','notice','notice_new','change_lv','change_draw','stock','on_order','stock_note','related_model',)
        elif type == 'CHECK':
            title = ('code', 'draw', 'name','row',)
        elif type == 'QTY':
            title = ('lv', 'name','num','row',)
        elif type in ('TEMP','REMOVE'):
            title=('lv', 'code', 'draw','name', 'num','other')

        try:
            wbook = load_workbook(file, read_only=True)
        except Exception as ex:
            rst['error']='文件读取失败：'+str(ex)
            return rst
        
        if type == 'CODE':
            names=wbook.sheetnames
        else:
            names = [wbook.active.title]

        for sname in names:  #读取多个工作表
            wsheet = wbook[sname]
            col=get_col(wsheet)
            if isinstance(col, dict):
                rst['col']=col
                read_item(wsheet)
            else:
                skip_sheet += wsheet.title +' : '+col+ ' ;'
                
        wbook.close()

        if skip_sheet:
            if type=='CODE':
                rst['skip'] = '跳过的工作表：'+skip_sheet
            else:
                rst['error'] = '缺少标题行：'+skip_sheet

        if item_error:
            item_error=[[1,x] for x in item_error]
            rst['itemerror'] = item_error
        
        if excel_bom:
            head=[]
            for key in title:
                head.append(key)
            excel_bom.insert(0, head)
            rst['bom'] = excel_bom
        rst['org_bom']=org_bom  #读取的原始内容

        return rst  #[0层次,1编码,2图号,3名称,4数量,5材料,6重量,7备注,8材料成本,9人工成本,10管理成本]

    def modify_excel(self,file,bom):
        '根据输入字典 { 行#列：{内容,备注,样式},  insert_col: ,}'
        fill_style={'blue':PatternFill('solid', fgColor='87CEEB'),
                    'red':PatternFill('solid', fgColor='FF4500'),
                    'green':PatternFill('solid', fgColor='00FF7F'),
                    'oringe':PatternFill('solid', fgColor='FFA500'),
                    'gray':PatternFill('solid', fgColor='A9A9A9'),
                    }
        rst={}
        try:
            wbook = load_workbook(file, read_only=False)
            wsheet = wbook.active
        except Exception as ex:
            rst['error']='文件读取失败：'+str(ex)
            return rst
        
        if 'insert_col' in bom:
            # 在数量列后面插入一列,inset会插入在前面,cols时列从1开始，而rows数组从0开始
            wsheet.insert_cols(bom['insert_col'])
        if 'modify' in bom:
            for key,val in bom['modify'].items():
                row=int(key.split('#')[0])
                col=int(key.split('#')[-1])
                if 'new' in val:
                    wsheet.cell(row,col).value = val['new']
                else:
                    wsheet.cell(row,col).value = val['old']

                if 'remark' in val:
                    wsheet.cell(row,col).comment = Comment(val['remark'],'check')
                if 'style' in val:
                    wsheet.cell(row,col).fill = fill_style[val['style']]

        try:
            wbook.save(file)
        except Exception as ex:
            rst['error'] = 'EXCEL文件写入失败：' + str(ex)        
        finally:
            wbook.close()

        return rst


    def bom_add_cost(self,bom):
        costbom={}
        all_cost,old_cost=self.load_cost_db()
        for item in bom:
            if item[1] in all_cost:
                cm = round(all_cost[item[1]][0],2)
                cl = round(all_cost[item[1]][1],2)
                ce = round(all_cost[item[1]][2],2)
                tot = round(all_cost[item[1]][3],2)
                cost_day = all_cost[item[1]][4]
            else:
                cm = cl = ce = tot = cost_day = 0
            
            iid=item[-1]
            if isinstance(item[5],(int,float)):
                tot_cost=round(tot * item[5], 2)
            else:
                tot_cost='' 
            item = tuple(item[:6]) + (cm, cl, ce, tot, tot_cost, cost_day,iid)          
            costbom[iid]=tuple('' if x == 0 else x for x in item)
        return costbom

    def recalc_tree_cost(self,cost_bom):
        '对装配体的子零件成本进行累加,如果和原装配体成本不同，则标示出'    

        tot_cm = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}   #层次材料汇总
        tot_cl = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}     #层次人工汇总
        tot_ce = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}     #层次管理费用汇总
        tot_ct = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0}     #层次总成本汇总
        #tot = {1: [0, 0, 0, 0], 2: [0, 0, 0, 0], 3: [0, 0, 0, 0], 4: [0, 0, 0, 0], 5: [0, 0, 0, 0], 6: [0, 0, 0, 0], 7: [0, 0, 0, 0],}

        complete_lv = {1: 1, 2: 1, 3: 1, 4: 1, 5: 1, 6: 1, 7: 1}
        for n, item in enumerate(cost_bom):
            item=list(item)
            for i in (6,7,8,9):
                if not item[i]:
                    item[i]=0
            item=tuple(item)

            if (item[1] == '' or item[1][:3] in ('24R', '28R') or item[9] == 0 or item[0]==1) and tot_ct[item[0] + 1] != 0:
                cm = round(tot_cm[item[0] + 1], 1)
                cl = round(tot_cl[item[0] + 1], 1)
                ce = round(tot_ce[item[0] + 1], 1)
                ct = round(tot_ct[item[0] + 1], 1)
                ct_t = round(ct * item[5], 1)
                #m = len(cost_bom) - n
                ss = ''
                if complete_lv[item[0] + 1] == 0:
                    ss = ' * '

                if abs(item[9] - ct) > 2:  # 原成本存在时和子件计算成本进行比较
                    old=ss+'{0}导入成本:{1} (材{2},人{3},费{4})'.format(item[11],str(item[9]),str(item[6]),str(item[7]),str(item[8]),)
                    cost_bom[n] = item[:6] + (cm, cl, ce, ct, ct_t) + (old,)+(item[-1],)
            else:
                cm = item[6]
                cl = item[7]
                ce = item[8]
                ct = item[9]

            if ct and isinstance(item[4],(int,float)):
                tot_cm[item[0]] += cm * item[4]
                tot_cl[item[0]] += cl * item[4]
                tot_ce[item[0]] += ce * item[4]
                tot_ct[item[0]] += ct * item[4]
            else:
                complete_lv[item[0]] = 0

            for i in range(item[0] + 1, 8):
                tot_cm[i] = 0
                tot_cl[i] = 0
                tot_ce[i] = 0
                tot_ct[i] = 0
                complete_lv[i] = 1

        bom={}
        for item in cost_bom[::-1]:
            bom[item[-1]]=tuple('' if x == 0 else x for x in item)

        return bom

    def scan_path(self, filedir, type=('.SLDDRW', '.DWG')):
        '扫描指定路径下图纸格式的文件，返回列表'
        def get_file(fdir):
            files = [x for x in listdir(fdir)]    #列出当前目录下所有内容
            patchs = [path.join(fdir, x) for x in files]  #拼接出当前目录下所有路径
            file_path = []
            for item in patchs:
                if path.isfile(item):
                    fname = path.basename(item).upper()
                    lname=path.splitext(item)[1].upper()
                    m_time=datetime.fromtimestamp(path.getmtime(item))
                    m_time=str(m_time.strftime('%Y-%m-%d  %H:%M'))
                    if lname in type:
                        name=fname.replace(lname,'')
                        file_path.append([name,item,m_time]) #后续还要添加产品信息，所以要用List
                elif path.isdir(item):
                    get_file(item)
            return file_path

        rst = {}
        drawpath=[]
        for root, dir, files in walk(filedir):
            for file in files:
                fpath = path.join(root, file)
                fname,lname=path.splitext(file.upper())
                m_time=datetime.fromtimestamp(path.getmtime(fpath))
                m_time=str(m_time.strftime('%Y-%m-%d  %H:%M'))
                if lname in type:
                    drawpath.append([fname,fpath,m_time])
        
        if drawpath:
            rst['path']=drawpath
        elif 'error' not in rst:
            rst['error'] = '选择的目录内没有图纸文件,请重新选择'

        return rst

    def read_json_date(self,filename):  # 从现有文件读取数据
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            print(filename+' 数据库读取失败')
            return {}

    def insert_db(self,new, sheet):
        '将信息写入数据库,有唯一项相同的进行替换。new为列表嵌套元组形式：[(),()]'
        cur = self.conn.cursor()
        table = {
            'batchBOM': ' VALUES (?,?,?)',
            'batchCODE': ' VALUES (?,?,?,?,?,?,?)',
            'root': ' VALUES (?,?,?)',
            'cost': ' VALUES (?,?,?,?,?,?,?)',
            'designBOM': ' VALUES (?,?,?,?)',
            'designCODE': ' VALUES (?,?,?,?,?,?,?)',
            'drawPATH': ' VALUES (?,?,?,?)',
            'design_change_item': ' VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)',
        }

        sql = 'INSERT OR REPLACE INTO ' + sheet + table[sheet]
        cur.executemany(sql, new)
        cur.close()

    def remove_db(self, del1='', col1='code', del2='', col2='', sheet=''):
        '从数据库中删除指定编码的物料,均为精确查找in'
        cur = self.conn.cursor()
        sql = 'DELETE FROM ' + sheet + ' WHERE ' + \
            col1 + ' in ' + str(del1+('deltet',))
        if col2 and del2:
            sql += 'AND ' + col2 + ' in ' + str(del2+('deltet',))

        cur.execute(sql)
        cur.close()

    def set_old_item(self,olditem, table):
        '对相应编码物料的remark设置为old'
        cur = self.conn.cursor()
        sql = 'UPDATE ' + table + \
            ' SET remark =\'old\' WHERE code in ' + str(olditem)
        #sql = r'UPDATE '+table + ' SET datetime = replace(datetime," 导入","")'   对字段进行替换
        cur.execute(sql)
        cur.close()

    def find_db(self,col, item, sheet):
        '从数据库查询物料,当item为字符串时，用Like-模糊查找；当item是元组时(至少2个元素)，用in-精确匹配'
        cur = self.conn.cursor()
        if item == '*':
            sql = 'SELECT * FROM ' + sheet
        else:
            sql = "PRAGMA table_info({0})".format(sheet)
            cur.execute(sql)
            cols = tuple(x[1] for x in cur.fetchall())  # 得到数据表所有列名
            if col == 'ALL':
                col = ''
                for m in cols[:-1]:
                    col += m+'||'
                col += cols[-1]
            elif col not in cols:
                return

            if isinstance(item, str):
                item = ' LIKE ' + '\'%' + item + '%\''
                #百分号（%）代表零个、一个或多个数字或字符。下划线（_）代表一个单一的数字或字符
            elif isinstance(item, tuple):
                item = ' IN '+str(item+('sdsdwewe@#',))
            else:
                return

            sql = 'SELECT * FROM ' + sheet + ' WHERE ' + col + item

        cur.execute(sql)
        rst = cur.fetchall()
        cur.close()
        return rst

    def db_command(self,s):
        if s in ('commit','ok'):
            self.conn.commit()
        elif s in ('rollback','back'):
            self.conn.rollback()

    def view_changed_cost(self,):
        '从数据库查询有变动过的成本，按编码排序'
        cost_changed = []
        code = ''
        for item in self.old_cost:  # cost库：[0编码，1材料，2人工，3费用，4总，5日期，6备注]
            if item[0] != code:
                code = item[0]
                cd, draw, name = self.get_code_info(code)[:3]
                cost_changed.append((1, code, draw, name) + self.all_cost[item[0]])
                cost_changed.append((2, code, draw, name) + item[1:])
            elif item[0] == code:
                cd, draw, name = self.get_code_info(code)[:3]
                cost_changed.append((2, code, draw, name) + item[1:])

        return cost_changed

    def save_to_excel(self,file, bom, title='', col={}):
        '把treeview内容保存到excel文件'
        if not file:
            return
        if not bom:
            return
        wb = Workbook()
        ws = wb.active

        ws.append([title])  # 添加标题

        colname = col.get('name', "")
        colwidth = col.get('width', "")

        if colname:
            ws.append(colname)  # 添加列名
        if colwidth:  # 设置列宽
            for n, wid in enumerate(colwidth):
                ws.column_dimensions[chr(65 + n)].width = wid//7
        #font_title = Font(u'微软雅黑', size=11,)  #设置字体样式

        order_num = [0, 0, 0, 0, 0, 0, 0, 0, 0]
        lv = {}
        for key in bom:
            i = key[0]
            order_num[i + 1:8] = 0, 0, 0, 0, 0, 0, 0, 0  # 物料序号
            order_num[i] += 1

            if i == 1:
                lv[i] = str(order_num[i])
            else:
                lv[i] = lv[i-1]+'.'+str(order_num[i])

            #level = tuple(order_num[key[0]] if i == key[0]-1 else '' for i in range(0, lv_col))
            ws.append((lv[i],)+tuple(key[1:]))

        rst = {}
        try:
            wb.save(file)
        except Exception as ex:
            rst['error'] = '文件保存错误: '+str(ex)
        return rst

    def download_template(self,file):
        '导出excel模板'
        wb = Workbook()
        ws = wb.active
        ws.title = '导入设计BOM模板'
        ws.append(('导入设计BOM时列名称规范，可以缺少列，但不能重复，层次列数量可以变化；BOM层次的第一行必须为ROOT',))
        ws.append(('层次', '层次', '层次', '层次', '层次', '编码',
                '图号', '名称', '数量', '材料', '单重', '总重', '备注'))
        ws.append(('ROOT', '', '', '', ''))

        ws1 = wb.create_sheet("导入ERP-BOM模板")
        ws1.append(('级别', '子件编码', '子件名称', '使用数量'))
        ws1.append(('ROOT', '', '', '', '', 'BOM层次的第一行必须为ROOT',))

        ws2 = wb.create_sheet("导入成本模板")
        ws2.append(('编码', '名称', '材料成本', '人工成本', '费用成本'))

        ws3 = wb.create_sheet("导入设计更改模板")
        ws3.append(('序号', '零部件图号', '零部件名称', '更改前编码', '更改后编码', '更改前说明', '更改后说明',
                '更改类别', '更改方式', '库存', '在途', '已制品处理建议', '涉及机型', '更改时间', '更改单号', '文件名'))

        rst = {}
        try:
            wb.save(file)
        except Exception as ex:
            rst['error'] = '文件保存错误: '+str(ex)
        return rst

    def creat_db(self,):
        "数据库初始化检查,如果没有则创建对应的表"
        cur = self.conn.cursor()
        batchBOM = "CREATE TABLE IF NOT EXISTS batchBOM (\
                    code VARCHAR, \
                    child VARCHAR, \
                    qty INTEGER NOT NULL, \
                    PRIMARY KEY(code, child)); "

        batchCODE = "CREATE TABLE IF NOT EXISTS batchCODE (\
                code VARCHAR PRIMARY KEY,\
                draw VARCHAR,\
                name VARCHAR, \
                metal VARCHAR, \
                weight VARCHAR, \
                remark VARCHAR, \
                time VARCHAR DEFAULT ''); "

        root = "CREATE TABLE IF NOT EXISTS root (\
                code VARCHAR, \
                name VARCHAR, \
                remark VARCHAR, \
                PRIMARY KEY(code, remark));"

        cost = "CREATE TABLE IF NOT EXISTS cost (\
                code VARCHAR,\
                cost_mt INTEGER,\
                cost_lb INTEGER, \
                cost_exp INTEGER, \
                cost_tot INTEGER, \
                datetime VARCHAR,\
                remark VARCHAR DEFAULT NULL); "

        # 插入默认时间:datetime VARCHAR DEFAULT (datetime('now','localtime'))

        designBOM = "CREATE TABLE IF NOT EXISTS designBOM (\
                    lv VARCHAR, \
                    code VARCHAR,\
                    qty INTEGER NOT NULL,\
                    root VARCHAR,\
                    PRIMARY KEY(lv,root));"

        designCODE = "CREATE TABLE IF NOT EXISTS designCODE (\
                    code VARCHAR PRIMARY KEY,\
                    draw VARCHAR,\
                    name VARCHAR,\
                    metal VARCHAR,\
                    weight INTEGER,\
                    remark INTEGER,\
                    time VARCHAR); "

        drawPATH = "CREATE TABLE IF NOT EXISTS drawPATH (\
                file VARCHAR,\
                path VARCHAR PRIMARY KEY,\
                mtime VARCHAR,\
                root VARCHAR); "
        design_change_item = "CREATE TABLE IF NOT EXISTS design_change_item (\
                        sn VARCHAR,\
                        draw VARCHAR,\
                        name VARCHAR,\
                        code_old VARCHAR,\
                        code_new VARCHAR,\
                        notice VARCHAR,\
                        notice_new VARCHAR,\
                        change_level VARCHAR,\
                        change_draw VARCHAR,\
                        stock VARCHAR,\
                        on_order VARCHAR,\
                        stock_clear VARCHAR,\
                        model VARCHAR,\
                        file VARCHAR,\
                        PRIMARY KEY(sn, file)); "

        for sql in (batchBOM, batchCODE, root, cost, designBOM, designCODE, drawPATH, design_change_item):
            try:
                cur.execute(sql)
            except Exception as ex:
                print("建表时出现如下异常:", ex)

        self.conn.commit()
        cur.close()
    def load_db(self,ss):
        cur = self.conn.cursor()
        cur.execute(ss)
        rst=list(cur.fetchall())
        cur.close()
        return rst

    def load_batch_db(self,):
        '读取数据库中小批物料和BOM信息'
        all_batch_code = {}
        all_batch_bom = {}
        root={}
        
        cur = self.conn.cursor()
        cur.execute('SELECT * FROM batchCODE')
        for item in cur.fetchall():
            all_batch_code[item[0]] = tuple(item)

        cur.execute('SELECT * FROM batchBOM')
        for item in cur.fetchall():
            if item[0] not in all_batch_bom:
                all_batch_bom[item[0]] = [(item[1], item[2])]
            else:
                all_batch_bom[item[0]].append((item[1], item[2]))
        cur.close()

        for key in all_batch_bom:
            if re.match(r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})', key):
                info = self.get_code_info(key)
                root[key] = info[2]
        
        return all_batch_code, all_batch_bom,root

    def load_cost_db(self,):  # 读取成本信息
        all_cost = {}
        old_cost=[]
        cur = self.conn.cursor()
        cur.execute('SELECT * FROM cost')
        for item in cur.fetchall():  # cost库：[0编码，1材料，2人工，3费用，4总，5日期，6备注]
            if not item[-1]:  #最后一位是备注，'old'代表旧价格
                all_cost[item[0]] = item[1:6]
            elif item[-1]=='old':
                old_cost.append(item[:6])
        old_cost.sort(key=lambda x: x[0])
        cur.close()
        return all_cost,old_cost

    def load_designCODE(self,):
        all_design_code = {}
        #cur = self.conn.cursor()
        #cur.execute('SELECT * FROM designCODE')
        rst=self.load_db('SELECT * FROM designCODE')
        for item in rst:
            all_design_code[item[0]] = tuple(item)

        return all_design_code

    def load_designBOM(self,):  # 读取设计BOM
        def sort_by_sn(item):
            '先将序号转为列表，再转为数字格式，再进行排序'
            sn = item[0].split('.')
            return [int(x) for x in sn]

        design_bom = []
        root = {}
        cur = self.conn.cursor()
        cur.execute('SELECT * FROM designBOM')
        design = {}
        for item in cur.fetchall():  # 先全部读取
            if item[0] == 'ROOT':
                item = ('1',)+item[1:]

            if item[-1] in design:
                design[item[-1]].append(item[:-1])
            else:
                design[item[-1]] = [item[:-1]]
        cur.close()

        for key in design:  # 按序号进行排序
            design[key].sort(key=sort_by_sn)

        lv_num = {0: 1, 1: 1}
        for val in design.values():
            for item in val:
                code, draw, name, metal, weight, remark = self.get_code_info(
                    item[1])
                num = item[2]
                lv = item[0].count('.')+1
                lv_num[lv] = lv_num[lv-1]*num
                design_bom.append((lv, code, draw, name, num,
                                lv_num[lv], metal, weight, remark))
        
        for key in design_bom:
            if key[0] == 1 and re.match(r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})', key[1]):
                root[key[1]] = key[3]

        return design_bom,root

    def load_design_change(self,):  # 读取设计更改信息
        def sort_by_sn(item):
            '先将序号转为列表，再转为数字格式，再进行排序'
            if item[1] == 'ROOT':
                return [0]
            else:
                sn = item[1].split('.')
                return [int(x) for x in sn]
        design_change_bom = []
        root={}
        cur = self.conn.cursor()
        cur.execute('SELECT * FROM design_change_item')
        change_bom = {}
        design_change_bom.clear()
        for item in cur.fetchall():
            file = item[-1]
            sn = item[0]
            if sn == 'ROOT':
                lv = 1
            else:
                lv = sn.count('.')+2

            if file in change_bom:
                change_bom[file].append((lv,)+item[:])
            else:
                change_bom[file] = [(lv,)+item[:]]
        cur.close()

        for key in change_bom:  # 按序号进行排序
            change_bom[key].sort(key=sort_by_sn)
            root[key]=''
        for val in change_bom.values():
            for item in val:
                design_change_bom.append(item)
 
        return design_change_bom,root

    def load_drawpath_db(self,):  # 读取图纸
        draw = {}
        root={}
        cur = self.conn.cursor()
        cur.execute('SELECT * FROM drawPATH')
        for item in cur.fetchall():
            if item[0] == 'ROOT':
                root[item[-1]] = item[1]
                continue
            elif item[-1] not in root:    #防止ROOT标志行丢失导致目录缺少
                root[item[-1]]=''

            if item[-1] not in draw:
                draw[item[-1]] = [item[:-1]]                
            else:
                draw[item[-1]].append(item[:-1])
        cur.close()

        return draw,root

    def get_batchRoot(self,):
        root={}
        for key in self.all_batch_bom:
            if re.match(r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})', key):
                info = self.get_code_info(key)
                root[key] = info[2]
        return root

    def get_designRoot(self,):
        root={}
        for key in self.all_design_bom:
            if key[0]==1 and re.match(r'(:?C|E|N|YF|EXP|CS)\d{2}(:?-|D|\d{2})', key[1]):                
                root[key[1]] = key[3]
        return root
    def get_designChangeRoot(self,):
        root={}
        rst = self.find_db('sn', ('ROOT',), 'design_change_item')
        for item in rst:
            root[item[2]] = ''
        return root
    def get_drawPathRoot(self,):
        root={}
        rst = self.find_db('file', ('ROOT',), 'drawPATH')
        for item in rst:
            root[item[-1]] = item[1]
        return root
    def get_root(self):
        return self.all_root

if __name__=='__main__':
    dataHandler()


